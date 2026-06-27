import openpyxl
from django.shortcuts import render
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import secrets
import string
from administration_app.pagination import ListPagination
from .serializers import *
from superadmin_app.models import *
from rest_framework.permissions import AllowAny , IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
import random
from datetime import datetime, timedelta
from django.utils import timezone
from django.core.mail import send_mail
import pyotp
# from .utils import decrypt_request_payload
from rest_framework_simplejwt.exceptions import TokenError
from collections import defaultdict
from io import BytesIO
from openpyxl.styles import Font
from openpyxl import Workbook
from django.db.models import Count
from superadmin_app.utils.security_utils import *
from django.db import transaction
from superadmin_app.authentication import *
from administration_app.pagination import ListPagination
from rest_framework.pagination import PageNumberPagination
from django.db.models import Sum
from django.db.models.functions import TruncMonth
from django.db.models import F

# Create your views here.

# class TokenRefreshAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         refresh_token = request.data.get("refresh")

#         if not refresh_token:

#             return Response({
#                 "status": False,
#                 "message": "Refresh token is required"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         try:

#             refresh = RefreshToken(refresh_token)

#             user_id = refresh.payload.get("user_id")

#             user = Profiles.objects.filter(
#                 id=user_id
#             ).first()

#             if not user:

#                 return Response({
#                     "status": False,
#                     "message": "User not found"
#                 }, status=status.HTTP_404_NOT_FOUND)

#             # =====================================
#             # TOKEN VERSION VALIDATION
#             # =====================================

#             token_version = refresh.payload.get(
#                 "token_version",
#                 0
#             )

#             if token_version != user.token_version:

#                 return Response({
#                     "status": False,
#                     "message":
#                     "Session expired. Please login again."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             access_token = refresh.access_token

#             # =====================================
#             # CUSTOM CLAIMS
#             # =====================================

#             access_token["profile_id"] = user.id
#             access_token["role"] = user.role
#             access_token["email"] = user.email
#             access_token["usersid"] = user.user_id
#             access_token["category"] = (
#                 user.category.name
#                 if user.category
#                 else None
#             )

#             access_token["token_version"] = (
#                 user.token_version
#             )

#             return Response({

#                 "status": True,
#                 "message":
#                 "Token refreshed successfully",

#                 "tokens": {
#                     "access": str(access_token),
#                     "refresh": str(refresh)
#                 }

#             }, status=status.HTTP_200_OK)

#         except TokenError:

#             return Response({
#                 "status": False,
#                 "message":
#                 "Token is invalid or expired"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         except Exception as e:

#             print(
#                 "TOKEN REFRESH ERROR =",
#                 str(e)
#             )

#             return Response({
#                 "status": False,
#                 "message":
#                 "Unable to refresh token."
#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# Create your views here.
# class TokenRefreshAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         refresh_token = request.data.get("refresh")

#         if not refresh_token:

#             return Response({
#                 "status": False,
#                 "message": "Refresh token is required"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         try:

#             refresh = RefreshToken(refresh_token)

#             user_id = refresh.payload.get("user_id")

#             user = Profiles.objects.filter(
#                 id=user_id
#             ).first()

#             if not user:

#                 return Response({
#                     "status": False,
#                     "message": "User not found"
#                 }, status=status.HTTP_404_NOT_FOUND)

#             # ============================
#             # SESSION VALIDATION
#             # ============================

#             session = UserSession.objects.filter(
#                 user=user,
#                 refresh_jti=refresh["jti"],
#                 is_active=True
#             ).first()

#             if not session:

#                 return Response({
#                     "status": False,
#                     "message": "Session expired or signed out."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             # ============================
#             # TOKEN VERSION VALIDATION
#             # ============================

#             token_version = refresh.payload.get(
#                 "token_version",
#                 0
#             )

#             if token_version != user.token_version:

#                 return Response({
#                     "status": False,
#                     "message":
#                     "Session expired. Please login again."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             # ============================
#             # CREATE NEW ACCESS TOKEN
#             # ============================

#             access_token = refresh.access_token

#             access_token["profile_id"] = user.id
#             access_token["role"] = user.role
#             access_token["email"] = user.email
#             access_token["usersid"] = user.user_id
#             access_token["category"] = (
#                 user.category.name
#                 if user.category
#                 else None
#             )

#             access_token["token_version"] = (
#                 user.token_version
#             )

#             # ============================
#             # UPDATE SESSION ACTIVITY
#             # ============================

#             session.last_activity = timezone.now()

#             session.save(
#                 update_fields=["last_activity"]
#             )

#             return Response({

#                 "status": True,

#                 "message":
#                 "Token refreshed successfully",

#                 "tokens": {

#                     "access": str(access_token),

#                     "refresh": str(refresh)

#                 }

#             }, status=status.HTTP_200_OK)

#         except TokenError:

#             return Response({

#                 "status": False,

#                 "message":
#                 "Token is invalid or expired"

#             }, status=status.HTTP_401_UNAUTHORIZED)

#         except Exception as e:

#             print(
#                 "TOKEN REFRESH ERROR =",
#                 str(e)
#             )

#             return Response({

#                 "status": False,

#                 "message":
#                 "Unable to refresh token."

#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# class LoginAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         # serializer = LoginSerializer(data=request.data)
#         print("REQUEST DATA:", request.data)

#         serializer = LoginSerializer(data=request.data)


#         print("SERIALIZER VALID:", serializer.is_valid())
#         print("SERIALIZER ERRORS:", serializer.errors)



#         if not serializer.is_valid():
#             return Response({
#                 "status": False,
#                 "errors": serializer.errors
#             }, status=status.HTTP_400_BAD_REQUEST)

#         user_id = serializer.validated_data.get("user_id")

#         password = serializer.validated_data.get("password")

#         category_id = serializer.validated_data.get("category_id")

#         user = Profiles.objects.filter(user_id=user_id).first()

#         user_agent = request.META.get('HTTP_USER_AGENT', '')
#         ip_address = get_client_ip(request)
#         browser = get_browser(user_agent)
#         device_name = get_device(user_agent)
#         location = get_location(ip_address)

#         if not user:

#             try:

#                 LoginHistory.objects.create(
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "User not found"
#             }, status=status.HTTP_404_NOT_FOUND)

#         if not user.check_password(password):

#             try:

#                 LoginHistory.objects.create(
#                     user=user,
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "Invalid password"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         # =========================================
#         # ASSIGN CATEGORY DURING LOGIN
#         # =========================================
#         if category_id:

#             try:
#                 category = Category.objects.get(id=category_id)

#                 user.category = category
#                 user.save()

#             except Category.DoesNotExist:

#                 return Response({
#                     "status": False,
#                     "message": "Invalid category_id"
#                 }, status=status.HTTP_400_BAD_REQUEST)

#         refresh = RefreshToken.for_user(user)
#         refresh['token_version'] = user.token_version

#         refresh_jti = refresh['jti']

#         access_token = refresh.access_token
#         session_jti = access_token['jti']

#         try:

#             print("ABOUT TO CREATE LOGIN HISTORY")

#             user_agent = request.META.get('HTTP_USER_AGENT', '')
#             ip_address = get_client_ip(request)
#             browser = get_browser(user_agent)
#             device_name = get_device(user_agent)
#             location = get_location(ip_address)

#             LoginHistory.objects.filter(
#                 user=user,
#                 logout_time__isnull=True
#             ).update(
#                 logout_time=timezone.now(),
#                 login_status='LOGOUT'
#             )

#             LoginHistory.objects.create(
#                 user=user,
#                 login_time=timezone.now(),
#                 login_status='SUCCESS',
#                 ip_address=ip_address,
#                 device_name=device_name,
#                 browser=browser,
#                 location=location,
#                 raw_user_agent=user_agent,
#                 )         
                

#             print("LOGIN HISTORY CREATED")

#         except Exception as e:

#             print(
#                 "LOGIN HISTORY ERROR =",
#                 str(e)
#             )

#         try:

#             print("ABOUT TO CREATE SESSION")

            

#             UserSession.objects.filter(
#                 user=user,
            
#                 is_active=True
#             ).update(is_active=False)

#             UserSession.objects.create(
#                 user=user,
#                 session_id=session_jti,
#                 refresh_jti=refresh_jti,
#                 token_version=user.token_version,
#                 device_name=device_name,
#                 browser=browser,
#                 ip_address=ip_address,
#                 location=location,
#                 is_active=True,
#                 expires_at=(timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])
#             )

            

#             print('ACTIVE SESSION CREATED FOR =', user.user_id)
#             print('EXPIRES AT =',timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])

#             print("SESSION CREATED")

#         except Exception as e:

#             print(
#                 "SESSION ERROR =",
#                 str(e)
#             )

#         refresh["profile_id"] = user.id
#         refresh["role"] = user.role
#         refresh["email"] = user.email
#         refresh["usersid"] = user.user_id
#         refresh["category"] = user.category.name if user.category else None
#         refresh["token_version"] = user.token_version
#         response  =  Response({
#             "status": True,
#             "message": "Login successful",

#             "user": {
#                 "id": user.id,
#                 "usersid": user.user_id,
#                 "role": user.role,
#                 "name": user.fullname or user.user_id,
#                 "category": user.category.name if user.category else None,
#             },

#             "tokens": {
#                 "access": str(access_token),
#                 "refresh": str(refresh),
#             }

#         }, status=status.HTTP_200_OK)

#         response.set_cookie(
#         key="access_token",
#         value=str(access_token),
#         httponly=True,
#         secure=False,
#         samesite="Lax",
#         max_age=15 * 60, 
#     )

#         response.set_cookie(
#             key="refresh_token",
#             value=str(refresh),
#             httponly=True,
#             secure=False,
#             samesite="Lax",
#              max_age=7 * 24 * 60 * 60,  # 7 days
#         )
#         return response


#   correct working api for refresh   
class TokenRefreshAPIView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request):

        refresh_token = request.data.get("refresh")

        if not refresh_token:
            refresh_token = request.COOKIES.get("refresh_token")

        if not refresh_token:
            return Response(
                {
                    "status": False,
                    "message": "Refresh token is required."
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:

            refresh = RefreshToken(refresh_token)

            user_id = refresh.get("user_id")

            user = Profiles.objects.filter(
                id=user_id
            ).first()

            if not user:
                return Response(
                    {
                        "status": False,
                        "message": "User not found."
                    },
                    status=status.HTTP_404_NOT_FOUND
                )

            # =====================================
            # CHECK TOKEN VERSION
            # =====================================

            token_version = refresh.get(
                "token_version",
                0
            )

            if token_version != user.token_version:

                response = Response(
                    {
                        "status": False,
                        "message": "Session expired. Please login again."
                    },
                    status=status.HTTP_401_UNAUTHORIZED
                )

                response.delete_cookie("access_token")
                response.delete_cookie("refresh_token")

                return response

            # =====================================
            # CHECK ACTIVE SESSION
            # =====================================

            session = UserSession.objects.filter(
                user=user,
                token_version=user.token_version,
                is_active=True
            ).first()

            if not session:

                response = Response(
                    {
                        "status": False,
                        "message": "Session expired. Please login again."
                    },
                    status=status.HTTP_401_UNAUTHORIZED
                )

                response.delete_cookie("access_token")
                response.delete_cookie("refresh_token")

                return response

            # =====================================
            # CHECK SESSION EXPIRY
            # =====================================

            if (
                session.expires_at and
                session.expires_at < timezone.now()
            ):

                session.is_active = False

                session.save(
                    update_fields=["is_active"]
                )

                response = Response(
                    {
                        "status": False,
                        "message": "Session expired."
                    },
                    status=status.HTTP_401_UNAUTHORIZED
                )

                response.delete_cookie("access_token")
                response.delete_cookie("refresh_token")

                return response

            session.last_activity = timezone.now()

            session.save(
                update_fields=["last_activity"]
            )

            # =====================================
            # CREATE NEW ACCESS TOKEN
            # =====================================

            access_token = refresh.access_token

            access_token["profile_id"] = user.id
            access_token["role"] = user.role
            access_token["email"] = user.email
            access_token["usersid"] = user.user_id
            access_token["category"] = (
                user.category.name
                if user.category else None
            )
            access_token["token_version"] = user.token_version
            # =====================================
            # UPDATE SESSION WITH NEW ACCESS TOKEN JTI
            # =====================================

            new_jti = access_token["jti"]

            print("OLD SESSION JTI :", session.session_id)
            print("NEW SESSION JTI :", new_jti)

            session.session_id = new_jti
            session.last_activity = timezone.now()

            session.save(
                update_fields=[
                    "session_id",
                    "last_activity",
                ]
            )

            response = Response(
                {
                    "status": True,
                    "message": "Token refreshed successfully.",
                    "tokens": {
                        "access": str(access_token),
                        "refresh": str(refresh)
                    }
                },
                status=status.HTTP_200_OK
            )

            response.set_cookie(
                key="access_token",
                value=str(access_token),
                httponly=True,
                secure=True,          # False only for local development
                samesite="None",
                # max_age=15 * 60,
                expires=timezone.now() + settings.SIMPLE_JWT["ACCESS_TOKEN_LIFETIME"],
            )

            response.set_cookie(
                key="refresh_token",
                value=str(refresh),
                httponly=True,
                secure=True,          # False only for local development
                samesite="None",
                # max_age=7 * 24 * 60 * 60,
                expires=timezone.now() + settings.SIMPLE_JWT["REFRESH_TOKEN_LIFETIME"],
            )

            return response

        except TokenError:

            response = Response(
                {
                    "status": False,
                    "message": "Refresh token is invalid or expired."
                },
                status=status.HTTP_401_UNAUTHORIZED
            )

            response.delete_cookie("access_token")
            response.delete_cookie("refresh_token")

            return response

        except Exception as e:

            print(e)

            return Response(
                {
                    "status": False,
                    "message": "Unable to refresh token."
                },
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



#  
# class TokenRefreshAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):
#         print("AUTH HEADER =", request.headers.get("Authorization"))
#         print("\n========== TOKEN REFRESH START ==========")

#         refresh_token = request.data.get("refresh")

#         print("REFRESH TOKEN FROM BODY =", refresh_token)

#         if not refresh_token:
#             refresh_token = request.COOKIES.get("refresh_token")
#             print("REFRESH TOKEN FROM COOKIE =", refresh_token)

#         if not refresh_token:
#             print("NO REFRESH TOKEN FOUND")
#             return Response({
#                 "status": False,
#                 "message": "Refresh token is required"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         try:

#             print("CREATING REFRESH TOKEN OBJECT")

#             refresh = RefreshToken(refresh_token)

#             print("REFRESH TOKEN CREATED SUCCESSFULLY")
#             print("TOKEN JTI =", refresh.get("jti"))
#             print("TOKEN USER ID =", refresh.get("user_id"))
#             print("TOKEN VERSION =", refresh.get("token_version"))
#             print("TOKEN EXP =", refresh.get("exp"))

#             user_id = refresh.payload.get("user_id")

#             print("FETCHING USER =", user_id)

#             user = Profiles.objects.filter(
#                 id=user_id
#             ).first()

#             session = UserSession.objects.filter(
#                 user=user,
#                 refresh_jti=refresh["jti"],
#                 is_active=True
#             ).first()

#             if not session:
#                 return Response({
#                     "status": False,
#                     "message": "Session expired or signed out."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             if not user:
#                 print("USER NOT FOUND")
#                 return Response({
#                     "status": False,
#                     "message": "User not found"
#                 }, status=status.HTTP_404_NOT_FOUND)

#             print("USER FOUND =", user.user_id)
#             print("DB TOKEN VERSION =", user.token_version)

#             token_version = refresh.payload.get(
#                 "token_version",
#                 0
#             )

#             print("TOKEN VERSION FROM JWT =", token_version)

#             if token_version != user.token_version:

#                 print(
#                     f"TOKEN VERSION MISMATCH | JWT={token_version} DB={user.token_version}"
#                 )

#                 return Response({
#                     "status": False,
#                     "message": "Session expired. Please login again."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             print("GENERATING NEW ACCESS TOKEN")

#             access_token = refresh.access_token

#             access_token["profile_id"] = user.id
#             access_token["role"] = user.role
#             access_token["email"] = user.email
#             access_token["usersid"] = user.user_id
#             access_token["category"] = (
#                 user.category.name
#                 if user.category else None
#             )
#             access_token["token_version"] = user.token_version

#             print("ACCESS TOKEN GENERATED")
#             print("ACCESS TOKEN JTI =", access_token.get("jti"))
#             print("ACCESS TOKEN EXP =", access_token.get("exp"))

#             session.session_id = access_token["jti"]
#             session.last_activity = timezone.now()

#             session.save(
#                 update_fields=[
#                     "session_id",
#                     "last_activity"
#                 ]
#             )

#             response = Response({

#                 "status": True,
#                 "message": "Token refreshed successfully",

#                 "tokens": {
#                     "access": str(access_token),
#                     "refresh": str(refresh)
#                 }

#             }, status=status.HTTP_200_OK)

#             print("SETTING ACCESS COOKIE")
#             print("SETTING REFRESH COOKIE")

#             response.set_cookie(
#                 key="access_token",
#                 value=str(access_token),
#                 httponly=True,
#                 secure=False,
#                 samesite="Lax",
#                 max_age=15 * 60,
#             )

#             response.set_cookie(
#                 key="refresh_token",
#                 value=str(refresh),
#                 httponly=True,
#                 secure=False,
#                 samesite="Lax",
#                 max_age=7 * 24 * 60 * 60,
#             )

#             print("TOKEN REFRESH SUCCESS")
#             print("========== TOKEN REFRESH END ==========\n")

#             return response

#         except TokenError as e:

#             print("TOKEN ERROR =", str(e))
#             print("========== TOKEN REFRESH FAILED ==========\n")

#             return Response({
#                 "status": False,
#                 "message": "Token is invalid or expired"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         except Exception as e:

#             print("UNEXPECTED TOKEN REFRESH ERROR =", str(e))
#             print("========== TOKEN REFRESH FAILED ==========\n")

#             return Response({
#                 "status": False,
#                 "message": "Unable to refresh token."
#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# # login
# class LoginAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         # serializer = LoginSerializer(data=request.data)
#         print("REQUEST DATA:", request.data)

#         serializer = LoginSerializer(data=request.data)


#         print("SERIALIZER VALID:", serializer.is_valid())
#         print("SERIALIZER ERRORS:", serializer.errors)



#         if not serializer.is_valid():
#             return Response({
#                 "status": False,
#                 "errors": serializer.errors
#             }, status=status.HTTP_400_BAD_REQUEST)

#         user_id = serializer.validated_data.get("user_id")

#         password = serializer.validated_data.get("password")

#         category_id = serializer.validated_data.get("category_id")

#         user = Profiles.objects.filter(user_id=user_id).first()

#         user_agent = request.META.get('HTTP_USER_AGENT', '')
#         ip_address = get_client_ip(request)
#         browser = get_browser(user_agent)
#         device_name = get_device(user_agent)
#         location = get_location(ip_address)

#         if not user:

#             try:

#                 LoginHistory.objects.create(
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "User not found"
#             }, status=status.HTTP_404_NOT_FOUND)

#         if not user.check_password(password):

#             try:

#                 LoginHistory.objects.create(
#                     user=user,
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "Invalid password"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         # =========================================
#         # ASSIGN CATEGORY DURING LOGIN
#         # =========================================
#         if category_id:

#             try:
#                 category = Category.objects.get(id=category_id)

#                 user.category = category
#                 user.save()

#             except Category.DoesNotExist:

#                 return Response({
#                     "status": False,
#                     "message": "Invalid category_id"
#                 }, status=status.HTTP_400_BAD_REQUEST)

#         refresh = RefreshToken.for_user(user)
#         refresh['token_version'] = user.token_version

#         refresh_jti = refresh["jti"]

#         access_token = refresh.access_token

#         try:

#             print("ABOUT TO CREATE LOGIN HISTORY")

#             user_agent = request.META.get('HTTP_USER_AGENT', '')
#             ip_address = get_client_ip(request)
#             browser = get_browser(user_agent)
#             device_name = get_device(user_agent)
#             location = get_location(ip_address)

#             LoginHistory.objects.create(
#                 user=user,
#                 login_time=timezone.now(),
#                 login_status='SUCCESS',
#                 ip_address=ip_address,
#                 device_name=device_name,
#                 browser=browser,
#                 location=location,
#                 raw_user_agent=user_agent,
#                 )         
                

#             print("LOGIN HISTORY CREATED")

#         except Exception as e:

#             print(
#                 "LOGIN HISTORY ERROR =",
#                 str(e)
#             )

#         try:

#             print("ABOUT TO CREATE SESSION")

            

#             session_jti = access_token["jti"]

#             UserSession.objects.create(
#                 user=user,
#                 session_id=session_jti,
#                 refresh_jti=refresh_jti,
#                 token_version=user.token_version,
#                 device_name=device_name,
#                 browser=browser,
#                 ip_address=ip_address,
#                 location=location,
#                 is_active=True,
#                 expires_at=(timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])
#             )

            

#             print('ACTIVE SESSION CREATED FOR =', user.user_id)
#             print('EXPIRES AT =',timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])

#             print("SESSION CREATED")

#         except Exception as e:

#             print(
#                 "SESSION ERROR =",
#                 str(e)
#             )

#         refresh["profile_id"] = user.id
#         refresh["role"] = user.role
#         refresh["email"] = user.email
#         refresh["usersid"] = user.user_id
#         refresh["category"] = user.category.name if user.category else None
#         refresh["token_version"] = user.token_version
#         response =  Response({
#             "status": True,
#             "message": "Login successful",

#             "user": {
#                 "id": user.id,
#                 "usersid": user.user_id,
#                 "role": user.role,
#                 "name": user.fullname or user.user_id,
#                 "category": user.category.name if user.category else None,
#                 "permissions" : user.permissions.values_list("code", flat=True)
#             },

#             "tokens": {
#                 "access": str(access_token),
#                 "refresh": str(refresh),
#             }

#         }, status=status.HTTP_200_OK)

#         response.set_cookie(
#         key="access_token",
#         value=str(access_token),
#         httponly=True,
#         secure=False,
#         samesite="Lax",
#         max_age=15 * 60, 
#     )

#         response.set_cookie(
#             key="refresh_token",
#             value=str(refresh),
#             httponly=True,
#             secure=False,
#             samesite="Lax",
#              max_age=7 * 24 * 60 * 60,  # 7 days
#         )
#         return response        

# login  correct login
# class LoginAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         # serializer = LoginSerializer(data=request.data)
#         print("REQUEST DATA:", request.data)

#         serializer = LoginSerializer(data=request.data)


#         print("SERIALIZER VALID:", serializer.is_valid())
#         print("SERIALIZER ERRORS:", serializer.errors)



#         if not serializer.is_valid():
#             return Response({
#                 "status": False,
#                 "errors": serializer.errors
#             }, status=status.HTTP_400_BAD_REQUEST)

#         user_id = serializer.validated_data.get("user_id")

#         password = serializer.validated_data.get("password")

#         category_id = serializer.validated_data.get("category_id")

#         user = Profiles.objects.filter(user_id=user_id).first()

#         user_agent = request.META.get('HTTP_USER_AGENT', '')
#         ip_address = get_client_ip(request)
#         browser = get_browser(user_agent)
#         device_name = get_device(user_agent)
#         location = get_location(ip_address)

#         if not user:

#             try:

#                 LoginHistory.objects.create(
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "User not found"
#             }, status=status.HTTP_404_NOT_FOUND)

#         if not user.check_password(password):

#             try:

#                 LoginHistory.objects.create(
#                     user=user,
#                     login_time=timezone.now(),
#                     login_status='FAILED',
#                     ip_address=ip_address,
#                     browser=browser,
#                     device_name=device_name,
#                     location=location,
#                     raw_user_agent=user_agent
#                 )

#             except Exception as e:

#                 print('FAILED LOGIN HISTORY ERROR =', str(e))

#             return Response({
#                 "status": False,
#                 "message": "Invalid password"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         # =========================================
#         # ASSIGN CATEGORY DURING LOGIN
#         # =========================================
#         if category_id:

#             try:
#                 category = Category.objects.get(id=category_id)

#                 user.category = category
#                 user.save()

#             except Category.DoesNotExist:

#                 return Response({
#                     "status": False,
#                     "message": "Invalid category_id"
#                 }, status=status.HTTP_400_BAD_REQUEST)

#         refresh = RefreshToken.for_user(user)
#         refresh['token_version'] = user.token_version

#         access_token = refresh.access_token
#         session_jti = access_token['jti']

#         try:

#             print("ABOUT TO CREATE LOGIN HISTORY")

#             user_agent = request.META.get('HTTP_USER_AGENT', '')
#             ip_address = get_client_ip(request)
#             browser = get_browser(user_agent)
#             device_name = get_device(user_agent)
#             location = get_location(ip_address)

#             LoginHistory.objects.filter(
#                 user=user,
#                 logout_time__isnull=True
#             ).update(
#                 logout_time=timezone.now(),
#                 login_status='LOGOUT'
#             )

#             LoginHistory.objects.create(
#                 user=user,
#                 login_time=timezone.now(),
#                 login_status='SUCCESS',
#                 ip_address=ip_address,
#                 device_name=device_name,
#                 browser=browser,
#                 location=location,
#                 raw_user_agent=user_agent,
#                 )         
                

#             print("LOGIN HISTORY CREATED")

#         except Exception as e:

#             print(
#                 "LOGIN HISTORY ERROR =",
#                 str(e)
#             )

#         try:

#             print("ABOUT TO CREATE SESSION")

            

#             UserSession.objects.filter(
#                 user=user,
            
#                 is_active=True
#             ).update(is_active=False)

#             UserSession.objects.create(
#                 user=user,
#                 session_id=session_jti,
#                 token_version=user.token_version,
#                 device_name=device_name,
#                 browser=browser,
#                 ip_address=ip_address,
#                 location=location,
#                 is_active=True,
#                 expires_at=(timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])
#             )

            

#             print('ACTIVE SESSION CREATED FOR =', user.user_id)
#             print('EXPIRES AT =',timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])

#             print("SESSION CREATED")

#         except Exception as e:

#             print(
#                 "SESSION ERROR =",
#                 str(e)
#             )

#         refresh["profile_id"] = user.id
#         refresh["role"] = user.role
#         refresh["email"] = user.email
#         refresh["usersid"] = user.user_id
#         refresh["category"] = user.category.name if user.category else None
#         refresh["token_version"] = user.token_version
#         response =  Response({
#             "status": True,
#             "message": "Login successful",

#             "user": {
#                 "id": user.id,
#                 "usersid": user.user_id,
#                 "role": user.role,
#                 "name": user.fullname or user.user_id,
#                 "category": user.category.name if user.category else None,
#                 "permissions" : user.permissions.values_list("code", flat=True)
#             },

#             "tokens": {
#                 "access": str(access_token),
#                 "refresh": str(refresh),
#             }

#         }, status=status.HTTP_200_OK)

#         response.set_cookie(
#         key="access_token",
#         value=str(access_token),
#         httponly=True,
#         secure=False,
#         samesite="Lax",
#         max_age=15 * 60, 
#     )

#         response.set_cookie(
#             key="refresh_token",
#             value=str(refresh),
#             httponly=True,
#             secure=False,
#             samesite="Lax",
#              max_age=7 * 24 * 60 * 60,  # 7 days
#         )
#         return response
        



# Create your views here.

# class LoginAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):

#         # serializer = LoginSerializer(data=request.data)
#         print("REQUEST DATA:", request.data)

#         serializer = LoginSerializer(data=request.data)

#         print("SERIALIZER VALID:", serializer.is_valid())
#         print("SERIALIZER ERRORS:", serializer.errors)

#         if not serializer.is_valid():
#             return Response({
#                 "status": False,
#                 "errors": serializer.errors
#             }, status=status.HTTP_400_BAD_REQUEST)

#         user_id = serializer.validated_data.get("user_id")

#         password = serializer.validated_data.get("password")

#         category_id = serializer.validated_data.get("category_id")

#         user = Profiles.objects.filter(user_id=user_id).first()

#         if not user:
#             return Response({
#                 "status": False,
#                 "message": "User not found"
#             }, status=status.HTTP_404_NOT_FOUND)

#         if not user.check_password(password):
#             return Response({
#                 "status": False,
#                 "message": "Invalid password"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         # =========================================
#         # ASSIGN CATEGORY DURING LOGIN
#         # =========================================
#         if category_id:

#             try:
#                 category = Category.objects.get(id=category_id)

#                 user.category = category
#                 user.save()

#             except Category.DoesNotExist:

#                 return Response({
#                     "status": False,
#                     "message": "Invalid category_id"
#                 }, status=status.HTTP_400_BAD_REQUEST)

#         refresh = RefreshToken.for_user(user)

#         refresh["profile_id"] = user.id
#         refresh["role"] = user.role
#         refresh["email"] = user.email
#         refresh["usersid"] = user.user_id
#         refresh["category"] = user.category.name if user.category else None

#         return Response({
#             "status": True,
#             "message": "Login successful",

#             "user": {
#                 "id": user.id,
#                 "usersid": user.user_id,
#                 "role": user.role,
#                 "name": user.fullname,
#                 "category": user.category.name if user.category else None,
#                 "permissions" : user.permissions.values_list("code", flat=True)
#             },

#             "tokens": {
#                 "access": str(refresh.access_token),
#                 "refresh": str(refresh),
#             }

#         }, status=status.HTTP_200_OK)

# class TokenRefreshAPIView(APIView):

#     permission_classes = [AllowAny]

#     def post(self, request):
#         print("AUTH HEADER =", request.headers.get("Authorization"))
#         print("\n========== TOKEN REFRESH START ==========")

#         refresh_token = request.data.get("refresh")

#         print("REFRESH TOKEN FROM BODY =", refresh_token)

#         if not refresh_token:
#             refresh_token = request.COOKIES.get("refresh_token")
#             print("REFRESH TOKEN FROM COOKIE =", refresh_token)

#         if not refresh_token:
#             print("NO REFRESH TOKEN FOUND")
#             return Response({
#                 "status": False,
#                 "message": "Refresh token is required"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         try:

#             print("CREATING REFRESH TOKEN OBJECT")

#             refresh = RefreshToken(refresh_token)

#             print("REFRESH TOKEN CREATED SUCCESSFULLY")
#             print("TOKEN JTI =", refresh.get("jti"))
#             print("TOKEN USER ID =", refresh.get("user_id"))
#             print("TOKEN VERSION =", refresh.get("token_version"))
#             print("TOKEN EXP =", refresh.get("exp"))

#             user_id = refresh.payload.get("user_id")

#             print("FETCHING USER =", user_id)

#             user = Profiles.objects.filter(
#                 id=user_id
#             ).first()

#             session = UserSession.objects.filter(
#                 user=user,
#                 token_version=user.token_version,
#                 is_active=True
#             ).order_by("-created_at").first()

#             if not session:
#                 return Response({
#                     "status": False,
#                     "message": "Session expired or signed out."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             if not user:
#                 print("USER NOT FOUND")
#                 return Response({
#                     "status": False,
#                     "message": "User not found"
#                 }, status=status.HTTP_404_NOT_FOUND)

#             print("USER FOUND =", user.user_id)
#             print("DB TOKEN VERSION =", user.token_version)

#             token_version = refresh.payload.get(
#                 "token_version",
#                 0
#             )

#             print("TOKEN VERSION FROM JWT =", token_version)

#             if token_version != user.token_version:

#                 print(
#                     f"TOKEN VERSION MISMATCH | JWT={token_version} DB={user.token_version}"
#                 )

#                 return Response({
#                     "status": False,
#                     "message": "Session expired. Please login again."
#                 }, status=status.HTTP_401_UNAUTHORIZED)

#             print("GENERATING NEW ACCESS TOKEN")

#             access_token = refresh.access_token

#             access_token["profile_id"] = user.id
#             access_token["role"] = user.role
#             access_token["email"] = user.email
#             access_token["usersid"] = user.user_id
#             access_token["category"] = (
#                 user.category.name
#                 if user.category else None
#             )
#             access_token["token_version"] = user.token_version

#             print("ACCESS TOKEN GENERATED")
#             print("ACCESS TOKEN JTI =", access_token.get("jti"))
#             print("ACCESS TOKEN EXP =", access_token.get("exp"))

#             session.session_id = access_token["jti"]
#             session.last_activity = timezone.now()

#             session.save(
#                 update_fields=[
#                     "session_id",
#                     "last_activity"
#                 ]
#             )

#             response = Response({

#                 "status": True,
#                 "message": "Token refreshed successfully",

#                 "tokens": {
#                     "access": str(access_token),
#                     "refresh": str(refresh)
#                 }

#             }, status=status.HTTP_200_OK)

#             print("SETTING ACCESS COOKIE")
#             print("SETTING REFRESH COOKIE")

#             response.set_cookie(
#                 key="access_token",
#                 value=str(access_token),
#                 httponly=True,
#                 secure=False,
#                 samesite="Lax",
#                 max_age=15 * 60,
#             )

#             response.set_cookie(
#                 key="refresh_token",
#                 value=str(refresh),
#                 httponly=True,
#                 secure=False,
#                 samesite="Lax",
#                 max_age=7 * 24 * 60 * 60,
#             )

#             print("TOKEN REFRESH SUCCESS")
#             print("========== TOKEN REFRESH END ==========\n")

#             return response

#         except TokenError as e:

#             print("TOKEN ERROR =", str(e))
#             print("========== TOKEN REFRESH FAILED ==========\n")

#             return Response({
#                 "status": False,
#                 "message": "Token is invalid or expired"
#             }, status=status.HTTP_401_UNAUTHORIZED)

#         except Exception as e:

#             print("UNEXPECTED TOKEN REFRESH ERROR =", str(e))
#             print("========== TOKEN REFRESH FAILED ==========\n")

#             return Response({
#                 "status": False,
#                 "message": "Unable to refresh token."
#             }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

# login
class LoginAPIView(APIView):
    authentication_classes = []

    permission_classes = [AllowAny]

    def post(self, request):

        # serializer = LoginSerializer(data=request.data)
        print("REQUEST DATA:", request.data)

        serializer = LoginSerializer(data=request.data)


        print("SERIALIZER VALID:", serializer.is_valid())
        print("SERIALIZER ERRORS:", serializer.errors)



        if not serializer.is_valid():
            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        user_id = serializer.validated_data.get("user_id")

        password = serializer.validated_data.get("password")

        category_id = serializer.validated_data.get("category_id")

        user = Profiles.objects.filter(user_id=user_id).first()

        user_agent = request.META.get('HTTP_USER_AGENT', '')
        ip_address = get_client_ip(request)
        browser = get_browser(user_agent)
        device_name = get_device(user_agent)
        location = get_location(ip_address)

        if not user:

            try:

                LoginHistory.objects.create(
                    login_time=timezone.now(),
                    login_status='FAILED',
                    ip_address=ip_address,
                    browser=browser,
                    device_name=device_name,
                    location=location,
                    raw_user_agent=user_agent
                )

            except Exception as e:

                print('FAILED LOGIN HISTORY ERROR =', str(e))

            return Response({
                "status": False,
                "message": "User not found"
            }, status=status.HTTP_404_NOT_FOUND)

        if not user.check_password(password):

            try:

                LoginHistory.objects.create(
                    user=user,
                    login_time=timezone.now(),
                    login_status='FAILED',
                    ip_address=ip_address,
                    browser=browser,
                    device_name=device_name,
                    location=location,
                    raw_user_agent=user_agent
                )

            except Exception as e:

                print('FAILED LOGIN HISTORY ERROR =', str(e))

            return Response({
                "status": False,
                "message": "Invalid password"
            }, status=status.HTTP_401_UNAUTHORIZED)

        # =========================================
        # ASSIGN CATEGORY DURING LOGIN
        # =========================================
        if category_id:

            try:
                category = Category.objects.get(id=category_id)

                user.category = category
                user.save()

            except Category.DoesNotExist:

                return Response({
                    "status": False,
                    "message": "Invalid category_id"
                }, status=status.HTTP_400_BAD_REQUEST)

        refresh = RefreshToken.for_user(user)
        refresh["profile_id"] = user.id
        refresh["role"] = user.role
        refresh["email"] = user.email
        refresh["usersid"] = user.user_id
        refresh["category"] = user.category.name if user.category else None
        refresh["token_version"] = user.token_version
       




        access_token = refresh.access_token

        try:

            print("ABOUT TO CREATE LOGIN HISTORY")

            user_agent = request.META.get('HTTP_USER_AGENT', '')
            ip_address = get_client_ip(request)
            browser = get_browser(user_agent)
            device_name = get_device(user_agent)
            location = get_location(ip_address)

            LoginHistory.objects.create(
                user=user,
                login_time=timezone.now(),
                login_status='SUCCESS',
                ip_address=ip_address,
                device_name=device_name,
                browser=browser,
                location=location,
                raw_user_agent=user_agent,
                )         
                

            print("LOGIN HISTORY CREATED")

        except Exception as e:

            print(
                "LOGIN HISTORY ERROR =",
                str(e)
            )

        try:

            print("ABOUT TO CREATE SESSION")

            

            session_jti = access_token["jti"]

            UserSession.objects.create(
                user=user,
                session_id=session_jti,
                token_version=user.token_version,
                device_name=device_name,
                browser=browser,
                ip_address=ip_address,
                location=location,
                is_active=True,
                expires_at=(
                    timezone.now()
                    + settings.SIMPLE_JWT["REFRESH_TOKEN_LIFETIME"]
                )
            )

            

            print('ACTIVE SESSION CREATED FOR =', user.user_id)
            print('EXPIRES AT =',timezone.now() + settings.SIMPLE_JWT['REFRESH_TOKEN_LIFETIME'])

            print("SESSION CREATED")

        except Exception as e:

            print(
                "SESSION ERROR =",
                str(e)
            )

        # refresh["profile_id"] = user.id
        # refresh["role"] = user.role
        # refresh["email"] = user.email
        # refresh["usersid"] = user.user_id
        # refresh["category"] = user.category.name if user.category else None
        # refresh["token_version"] = user.token_version

        # access_token["profile_id"] = user.id
        # access_token["role"] = user.role
        # access_token["email"] = user.email
        # access_token["usersid"] = user.user_id
        # access_token["category"] = user.category.name if user.category else None
        # access_token["token_version"] = user.token_version

        response =  Response({
            "status": True,
            "message": "Login successful",

            "user": {
                "id": user.id,
                "usersid": user.user_id,
                "role": user.role,
                "name": user.fullname or user.user_id,
                "category": user.category.name if user.category else None,
                "permissions" : user.permissions.values_list("code", flat=True)
            },

            "tokens": {
                "access": str(access_token),
                "refresh": str(refresh),
            }

        }, status=status.HTTP_200_OK)

        response.set_cookie(
        key="access_token",
        value=str(access_token),
        httponly=True,
        # secure=True,
            secure=False,
        # samesite="None",
         samesite="Lax",
        expires=timezone.now() + settings.SIMPLE_JWT["ACCESS_TOKEN_LIFETIME"],
        # max_age=15 * 60, 
        
    )

        response.set_cookie(
            key="refresh_token",
            value=str(refresh),
            httponly=True,
            # secure=True,
                secure=False,
            # samesite="None",
             samesite="Lax",
            #  max_age=7 * 24 * 60 * 60,  # 7 days
            expires=timezone.now() + settings.SIMPLE_JWT["REFRESH_TOKEN_LIFETIME"],
        )
        return response    

# list category
class ListCategoryAPIView(APIView):

    permission_classes = [AllowAny]
    authentication_classes = [CookieOrHeaderJWTAuthentication]

    def get(self, request):

        categories = Category.objects.all()

        data = []

        for category in categories:
            data.append({
                "id": category.id,
                "name": category.name,
                "description": category.description
            })

        return Response({
            "status": True,
            "categories": data
        }, status=status.HTTP_200_OK)



# forgot password view
class ForgotPasswordAPIView(APIView):

    permission_classes = [AllowAny]

    def post(self, request):

        email = request.data.get("email")

        if not email:
            return Response({
                "status": False,
                "message": "Email is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        user = Profiles.objects.filter(
            email=email
        ).first()

        if not user:
            return Response({
                "status": False,
                "message": "User not found"
            }, status=status.HTTP_404_NOT_FOUND)

        OTPVerification.objects.filter(
            user=user,
            purpose='forgot_password',
            is_used=False
        ).update(
            is_used=True
        )

        secret = pyotp.random_base32()

        totp = pyotp.TOTP(
            secret,
            interval=300
        )

        otp = totp.now()

        OTPVerification.objects.create(
            user=user,
            secret=secret,
            otp=otp,
            purpose='forgot_password',
            expires_at=timezone.now() + timedelta(minutes=5)
        )

        send_mail(
            subject="Password Reset OTP",
            message=f"Your OTP is {otp}. It will expire in 5 minutes.",
            from_email=None,
            recipient_list=[email],
            fail_silently=False
        )

        return Response({
            "status": True,
            "message": "OTP sent successfully"
        })
    

# verify otp 
class VerifyOTPAPIView(APIView):

    permission_classes = [AllowAny]

    def post(self, request):

        email = request.data.get("email")
        otp = request.data.get("otp")

        if not email:
            return Response({
                "status": False,
                "message": "Email is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if not otp:
            return Response({
                "status": False,
                "message": "OTP is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        user = Profiles.objects.filter(email=email).first()

        if not user:
            return Response({
                "status": False,
                "message": "User not found"
            }, status=status.HTTP_404_NOT_FOUND)

        otp_record = OTPVerification.objects.filter(
            user=user,
            purpose='forgot_password',
            is_used=False
        ).order_by('-created_at').first()

        if not otp_record:
            return Response({
                "status": False,
                "message": "OTP not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Check expiry
        if timezone.now() > otp_record.expires_at:
            return Response({
                "status": False,
                "message": "OTP expired"
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check max attempts
        if otp_record.attempts >= otp_record.max_attempts:
            return Response({
                "status": False,
                "message": "Maximum OTP attempts exceeded"
            }, status=status.HTTP_400_BAD_REQUEST)

        # Increase attempts
        otp_record.attempts += 1
        otp_record.save(update_fields=["attempts"])

        # Verify OTP using pyotp
        totp = pyotp.TOTP(
            otp_record.secret,
            interval=300
        )

        if not totp.verify(str(otp)):
            return Response({
                "status": False,
                "message": "Invalid OTP",
                "remaining_attempts":
                    otp_record.max_attempts - otp_record.attempts
            }, status=status.HTTP_400_BAD_REQUEST)

        # Mark OTP as used
        otp_record.is_used = True
        otp_record.used_at = timezone.now()
        otp_record.save(update_fields=["is_used", "used_at"])

        return Response({
            "status": True,
            "message": "OTP verified successfully"
        }, status=status.HTTP_200_OK)
    

    # reset pasword section 
class ResetPasswordAPIView(APIView):

    permission_classes = [AllowAny]

    def post(self, request):

        email = request.data.get("email")
        new_password = request.data.get("new_password")
        confirm_password = request.data.get("confirm_password")

        if not email:
            return Response({
                "status": False,
                "message": "Email is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if not new_password:
            return Response({
                "status": False,
                "message": "New password is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if not confirm_password:
            return Response({
                "status": False,
                "message": "Confirm password is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({
                "status": False,
                "message": "Passwords do not match"
            }, status=status.HTTP_400_BAD_REQUEST)

        user = Profiles.objects.filter(email=email).first()

        if not user:
            return Response({
                "status": False,
                "message": "User not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Check whether OTP was verified
        verified_otp = OTPVerification.objects.filter(
            user=user,
            purpose="forgot_password",
            is_used=True
        ).order_by("-used_at").first()

        if not verified_otp:
            return Response({
                "status": False,
                "message": "OTP verification required before resetting password"
            }, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save()

        # Optional: delete or invalidate verified OTP records
        OTPVerification.objects.filter(
            user=user,
            purpose="forgot_password"
        ).delete()

        return Response({
            "status": True,
            "message": "Password reset successfully"
        }, status=status.HTTP_200_OK)

#resend otp 
class ResendForgotPasswordOTPAPIView(APIView):

    permission_classes = [AllowAny]

    def post(self, request):

        email = request.data.get("email")

        if not email:
            return Response(
                {
                    "status": False,
                    "message": "Email is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user = Profiles.objects.get(email=email)

        except Profiles.DoesNotExist:
            return Response(
                {
                    "status": False,
                    "message": "Email not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        OTPVerification.objects.filter(
            user=user,
            purpose="forgot_password",
            is_used=False
        ).update(is_used=True)

        # Generate Secret
        secret = pyotp.random_base32()

        # Generate OTP valid for 1 minute
        totp = pyotp.TOTP(
            secret,
            interval=300
        )

        otp = totp.now()

        OTPVerification.objects.create(
            user=user,
            otp=otp,
            secret=secret,
            purpose="forgot_password",
            is_used=False,
            attempts=0,
            expires_at=timezone.now() + timedelta(minutes=5)
        )

        send_mail(
            subject="OTP to Reset Your Password",
            message=f"Your OTP is {otp}. It is valid for 5 minutes.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[email],
            fail_silently=False
        )

        return Response(
            {
                "status": True,
                "message": "OTP resent successfully",
            },
            status=status.HTTP_200_OK
        )
    





#***************************Attendance management views for teachers *******************************
# kpi cards 
class AttendanceManagementKPICardsAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date_param = request.query_params.get("date")
        if date_param:
            try:
                selected_date = datetime.strptime(date_param, "%Y-%m-%d").date()
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Invalid date format. Use YYYY-MM-DD"
                }, status=status.HTTP_400_BAD_REQUEST)
        else:
            selected_date = timezone.now().date()

        # =======================
        # Students
        # =======================

        total_students = Profiles.objects.filter(
            role="Student"
        ).count()

        present_students = StudentAttendance.objects.filter(
            date=selected_date,
            status="Present"
        ).count()

        student_attendance_rate = (
            round((present_students / total_students) * 100, 2)
            if total_students > 0 else 0
        )

        # =======================
        # Teachers
        # =======================

        total_teachers = StaffManagementModel.objects.filter(
            designation="Teacher",
            is_teacher=True
        ).count()

        present_teachers = TeacherstaffAttendance.objects.filter(
            date=selected_date,
            status="Present",
            teacher__is_teacher=True
        ).count()

        absent_teachers = TeacherstaffAttendance.objects.filter(
            date=selected_date,
            status="Absent",
            teacher__is_teacher=True
        ).count()

        teacher_attendance_rate = (
            round((present_teachers / total_teachers) * 100, 2)
            if total_teachers > 0 else 0
        )

        # =======================
        # Staff
        # =======================

        total_staff = StaffManagementModel.objects.filter(
            is_teacher=False
        ).count()

        present_staff = TeacherstaffAttendance.objects.filter(
            date=selected_date,
            status="Present",
            teacher__is_teacher=False
        ).count()

        absent_staff = TeacherstaffAttendance.objects.filter(
            date=selected_date,
            status="Absent",
            teacher__is_teacher=False
        ).count()

        staff_attendance_rate = (
            round((present_staff / total_staff) * 100, 2)
            if total_staff > 0 else 0
        )

        return Response({
            "status": True,
            "message": "Attendance KPI cards data retrieved successfully",
            "data": [
        {
            "title": "Student Attendance Today",
            "present": present_students,
            "total": total_students,
            "count": f"{present_students}/{total_students}",
            "attendance_rate": student_attendance_rate
        },
        {
            "title": "Teachers Attendance Today",
            "present": present_teachers,
            "total": total_teachers,
            "count": f"{present_teachers}/{total_teachers}",
            "attendance_rate": teacher_attendance_rate
        },
        {
            "title": "Staff Attendance Today",
            "present": present_staff,
            "total": total_staff,
            "count": f"{present_staff}/{total_staff}",
            "attendance_rate": staff_attendance_rate
        }
    ]
})


# list teachers attendance based on date
# class TeacherAttendanceListAPIView(APIView):

#     permission_classes = [IsAuthenticated]

#     def get(self, request):

#         date = request.query_params.get("date")

#         if not date:
#             return Response({
#                 "status": False,
#                 "message": "Date query parameter is required in YYYY-MM-DD format"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         try:
#             selected_date = datetime.strptime(date, "%Y-%m-%d").date()
#         except ValueError:
#             return Response({
#                 "status": False,
#                 "message": "Invalid date format. Use YYYY-MM-DD"
#             }, status=status.HTTP_400_BAD_REQUEST)

#         teachers = StaffManagementModel.objects.filter(designation="Teacher",is_teacher=True)

#         attendance_records = TeacherstaffAttendance.objects.filter(
#             date=selected_date
#         )

#         attendance_dict = {
#             attendance.teacher_id: attendance
#             for attendance in attendance_records
#         }

#         data = []

#         for teacher in teachers:

#             attendance = attendance_dict.get(teacher.id)

#             data.append({
#                 "id" : attendance.id if attendance else None,

#                 "profile_id": teacher.id,
#                 "teacher_id": teacher.profiles.user_id,
#                 "teacher_name": teacher.staff_name,
#                 "teacher_department": teacher.department,
                
#                 "date": selected_date,
#                 "status": attendance.status if attendance else "Not Marked",
#                 "remarks": attendance.remarks if attendance else None,
#                 "checked_in_time": attendance.checked_in_time if attendance else None,
#                 "checked_out_time": attendance.checked_out_time if attendance else None,
#             })

#         # Pagination
#         paginator = ListPagination()

#         paginated_data = paginator.paginate_queryset(
#             data,
#             request
#         )

#         return paginator.get_paginated_response({
#             "status": True,
#             "message": "Teacher attendance records retrieved successfully",
#             "attendance": paginated_data
#         }, status=status.HTTP_200_OK)

class TeacherAttendanceListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            selected_date = datetime.strptime(
                date,
                "%Y-%m-%d"
            ).date()
        except ValueError:
            return Response({
                "status": False,
                "message": "Invalid date format. Use YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        teachers = StaffManagementModel.objects.filter(
            designation="Teacher",
            is_teacher=True
        ).select_related("profiles")

        attendance_records = TeacherstaffAttendance.objects.filter(
            date=selected_date
        )

        attendance_dict = {
            attendance.teacher_id: attendance
            for attendance in attendance_records
        }

        # Get levels handled by each teacher
        teacher_level_map = defaultdict(list)

        timetable_data = TeachersTimeTable.objects.select_related(
            "class_assigned"
        ).values(
            "teacher_id",
            "class_assigned__level"
        ).distinct()

        for item in timetable_data:

            level = item.get("class_assigned__level")

            if level and level not in teacher_level_map[item["teacher_id"]]:
                teacher_level_map[item["teacher_id"]].append(level)

        data = []

        for teacher in teachers:

            attendance = attendance_dict.get(teacher.id)

            data.append({
                "id": attendance.id if attendance else None,

                "profile_id": teacher.id,
                "teacher_id": teacher.profiles.user_id if teacher.profiles else None,
                "teacher_name": teacher.staff_name,
                "teacher_department": teacher.department,

                # Added level
                "level": teacher_level_map.get(teacher.id, []),

                "date": selected_date,
                "status": attendance.status if attendance else "Not Marked",
                "remarks": attendance.remarks if attendance else None,
                "checked_in_time": attendance.checked_in_time if attendance else None,
                "checked_out_time": attendance.checked_out_time if attendance else None,
            })

        paginator = ListPagination()

        paginated_data = paginator.paginate_queryset(
            data,
            request
        )

        return paginator.get_paginated_response({
            "status": True,
            "message": "Teacher attendance records retrieved successfully",
            "attendance": paginated_data
        })



# mark attendance for teachers 
# class MarkTeachersAttendance(APIView):

#     permission_classes = [IsAuthenticated]

#     def post(self, request):

#         serializer = MarkTeachersAttendanceSerializer(data=request.data)

#         if not serializer.is_valid():
#             return Response({
#                 "status": False,
#                 "errors": serializer.errors
#             }, status=status.HTTP_400_BAD_REQUEST)

#         teacher_id = serializer.validated_data.get("teacher_id")
#         date = serializer.validated_data.get("date")
#         status_attendance = serializer.validated_data.get("status")
#         remarks = serializer.validated_data.get("remarks", None)

#         checked_in_time = serializer.validated_data.get("checked_in_time", None)
#         checked_out_time = serializer.validated_data.get("checked_out_time", None)

#         teacher = StaffManagementModel.objects.filter(
#             id=teacher_id,
#             designation="Teacher",
#             is_teacher=True
#         ).first()

#         if not teacher:
#             return Response({
#                 "status": False,
#                 "message": "Teacher not found"
#             }, status=status.HTTP_404_NOT_FOUND)

#         attendance_record, created = TeacherstaffAttendance.objects.get_or_create(
#             teacher=teacher,
#             date=date
#         )

#         # Always update attendance status
#         attendance_record.status = status_attendance

#         # Update remarks only if provided
#         if remarks is not None:
#             attendance_record.remarks = remarks

#         # Update check-in time only if provided
#         if checked_in_time is not None:
#             attendance_record.checked_in_time = checked_in_time

#         # Update check-out time only if provided
#         if checked_out_time is not None:
#             attendance_record.checked_out_time = checked_out_time

#         # Optional: clear times when absent
#         if status_attendance == "Absent":
#             attendance_record.checked_in_time = None
#             attendance_record.checked_out_time = None

#         attendance_record.save()

#         return Response({
#             "status": True,
#             "message": (
#                 "Teacher attendance updated successfully"
#                 if not created
#                 else "Teacher attendance marked successfully"
#             ),
#             "attendance": {
#                 "attendance_id": attendance_record.id,
#                 "teacher_id": teacher.id,
#                 "teacher_name": teacher.staff_name,
#                 "teacher_department": teacher.department,
#                 "date": attendance_record.date,
#                 "status": attendance_record.status,
#                 "remarks": attendance_record.remarks,
#                 "checked_in_time": attendance_record.checked_in_time,
#                 "checked_out_time": attendance_record.checked_out_time,
#             }
#         }, status=status.HTTP_200_OK)

# mark attendance to teachers
class MarkTeachersAttendance(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = MarkTeachersAttendanceSerializer(
            data=request.data,
            partial=True
        )

        if not serializer.is_valid():
            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        teacher_id = serializer.validated_data.get("teacher_id")
        date = serializer.validated_data.get("date")

        if not teacher_id or not date:
            return Response({
                "status": False,
                "message": "teacher_id and date are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        teacher = StaffManagementModel.objects.filter(
            id=teacher_id,
            designation="Teacher",
            is_teacher=True
        ).first()

        if not teacher:
            return Response({
                "status": False,
                "message": "Teacher not found"
            }, status=status.HTTP_404_NOT_FOUND)

        attendance_record, created = TeacherstaffAttendance.objects.get_or_create(
            teacher=teacher,
            date=date,
            defaults={
        "is_staff": False
    })
        attendance_record.is_staff = False


        

        # Partial updates
        if "status" in serializer.validated_data:
            attendance_record.status = serializer.validated_data["status"]

            if serializer.validated_data["status"] == "Absent":
                attendance_record.checked_in_time = None
                attendance_record.checked_out_time = None

        if "remarks" in serializer.validated_data:
            attendance_record.remarks = serializer.validated_data["remarks"]

        if "checked_in_time" in serializer.validated_data:
            attendance_record.checked_in_time = serializer.validated_data["checked_in_time"]

        if "checked_out_time" in serializer.validated_data:
            attendance_record.checked_out_time = serializer.validated_data["checked_out_time"]

        attendance_record.save()

        return Response({
            "status": True,
            "message": (
                "Teacher attendance updated successfully"
                if not created
                else "Teacher attendance marked successfully"
            ),
            "attendance": {
                "attendance_id": attendance_record.id,
                "teacher_id": teacher.id,
                "teacher_name": teacher.staff_name,
                "teacher_department": teacher.department,
                "date": attendance_record.date,
                "status": attendance_record.status,
                "remarks": attendance_record.remarks,
                "checked_in_time": attendance_record.checked_in_time,
                "checked_out_time": attendance_record.checked_out_time,
            }
        }, status=status.HTTP_200_OK)
    


   # students attendance list 
class WebStudentAttendanceListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")
        class_id = request.query_params.get("class_id")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        classes = ClassModel.objects.all()

        if class_id:
            classes = classes.filter(id=class_id)

        response_data = []

        for cls in classes:

            total_students = StudentAcademicDetails.objects.filter(
                student_class=cls
            ).count()

            attendance_qs = StudentAttendance.objects.filter(
                students_class=cls,
                date=date
            )

            present_count = attendance_qs.filter(
                status="Present"
            ).count()

            absent_count = attendance_qs.filter(
                status="Absent"
            ).count()

            half_day_count = attendance_qs.filter(
                status="Half Day"
            ).count()

            taken_by = attendance_qs.first().taken_by if attendance_qs.exists() else None

            response_data.append({
                "class_id": cls.id,
                "class_code": getattr(cls, "class_id", ""),
                "class_name": getattr(cls, "class_name", ""),
                "section": getattr(cls, "section", ""),
                "total_students": total_students,
                "present": present_count,
                "absent": absent_count,
                "half_day": half_day_count,
                "attendance_taken_by": {
                    "id": taken_by.id if taken_by else None,
                    "name": taken_by.staff_name if taken_by else None
                } if taken_by else None,
                "incharge": {
                    "id": cls.class_teacher.id if hasattr(cls, "class_teacher") and cls.class_teacher else None,
                    "name": cls.class_teacher.staff_name if hasattr(cls, "class_teacher") and cls.class_teacher else None
                }
            })

        return Response({
            "status": True,
            "message": "Attendance list fetched successfully",
            "data": response_data
        }, status=status.HTTP_200_OK)
    




# list staff attendance based on date
class StaffAttendanceListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        staff_members = StaffManagementModel.objects.filter(is_teacher=False)

        attendance_records = TeacherstaffAttendance.objects.filter(
            date=date,
            teacher__is_teacher=False
        )

        attendance_dict = {
            attendance.teacher_id: attendance
            for attendance in attendance_records
        }

        data = []

        for staff in staff_members:

            attendance = attendance_dict.get(staff.id)

            data.append({
                "id" : attendance.id if attendance else None,

                "profile_id": staff.id,
                "staff_id": staff.profiles.user_id,
                "staff_name": staff.staff_name,
                "staff_department": staff.department,
                "staff_designation" :staff.designation,

                "date": date,
                "status": attendance.status if attendance else "Not Marked",
                "remarks": attendance.remarks if attendance else None,
                "checked_in_time": attendance.checked_in_time if attendance else None,
                "checked_out_time": attendance.checked_out_time if attendance else None,
            })

        # Pagination
        paginator = ListPagination()

        paginated_data = paginator.paginate_queryset(
            data,
            request
        )

        return paginator.get_paginated_response({
            "status": True,
            "message": "Staff attendance records retrieved successfully",
            "attendance": paginated_data
        })






  # mark staff attendance   
class MarkStaffAttendanceAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = MarkStaffAttendanceSerializer(
            data=request.data,
            partial=True
        )

        if not serializer.is_valid():
            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        staff_id = serializer.validated_data.get("staff_id")
        date = serializer.validated_data.get("date")

        if not staff_id or not date:
            return Response({
                "status": False,
                "message": "staff_id and date are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        teacher = StaffManagementModel.objects.filter(
            id=staff_id,
            is_teacher=False
        ).first()

        if not teacher:
            return Response({
                "status": False,
                "message": "Teacher not found"
            }, status=status.HTTP_404_NOT_FOUND)

        attendance_record, created = TeacherstaffAttendance.objects.get_or_create(
            teacher=teacher,
            date=date,
            defaults={
        "is_staff": True
    })
        attendance_record.is_staff = True


        

        # Partial updates
        if "status" in serializer.validated_data:
            attendance_record.status = serializer.validated_data["status"]

            if serializer.validated_data["status"] == "Absent":
                attendance_record.checked_in_time = None
                attendance_record.checked_out_time = None

        if "remarks" in serializer.validated_data:
            attendance_record.remarks = serializer.validated_data["remarks"]

        if "checked_in_time" in serializer.validated_data:
            attendance_record.checked_in_time = serializer.validated_data["checked_in_time"]

        if "checked_out_time" in serializer.validated_data:
            attendance_record.checked_out_time = serializer.validated_data["checked_out_time"]

        attendance_record.save()

        return Response({
            "status": True,
            "message": (
                "Staffs attendance updated successfully"
                if not created
                else "Staffs attendance marked successfully"
            ),
            "attendance": {
                "attendance_id": attendance_record.id,
                "staff_id": teacher.id,
                "staff_name": teacher.staff_name,
                "teacher_department": teacher.department,
                "date": attendance_record.date,
                "status": attendance_record.status,
                "remarks": attendance_record.remarks,
                "checked_in_time": attendance_record.checked_in_time,
                "checked_out_time": attendance_record.checked_out_time,
            }
        }, status=status.HTTP_200_OK)
    


# export  excel sheet for teachers attedance , students attendance and also for staffs attendance

# export teachers attendance excel sheet based on date , department 
class ExportExcelTeachersAttendanceAPI(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            selected_date = datetime.strptime(
                date,
                "%Y-%m-%d"
            ).date()

        except ValueError:
            return Response({
                "status": False,
                "message": "Invalid date format. Use YYYY-MM-DD"
            }, status=status.HTTP_400_BAD_REQUEST)

        teachers = StaffManagementModel.objects.filter(
            designation="Teacher",
            is_teacher=True
        ).select_related("profiles")

        attendance_records = TeacherstaffAttendance.objects.filter(
            date=selected_date
        )

        attendance_dict = {
            attendance.teacher_id: attendance
            for attendance in attendance_records
        }

        teacher_level_map = defaultdict(list)

        timetable_data = TeachersTimeTable.objects.select_related(
            "class_assigned"
        ).values(
            "teacher_id",
            "class_assigned__level"
        ).distinct()

        for item in timetable_data:

            level = item.get("class_assigned__level")

            if level and level not in teacher_level_map[item["teacher_id"]]:
                teacher_level_map[item["teacher_id"]].append(level)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Teacher Attendance"

        headers = [
            "Teacher ID",
            "Teacher Name",
            "Department",
            "Level",
            "Date",
            "Status",
            "Remarks",
            "Check In Time",
            "Check Out Time"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        row_num = 2

        for teacher in teachers:

            attendance = attendance_dict.get(teacher.id)

            levels = ", ".join(
                map(str, teacher_level_map.get(teacher.id, []))
            )

            worksheet.cell(
                row=row_num,
                column=1,
                value=teacher.profiles.user_id if teacher.profiles else ""
            )

            worksheet.cell(
                row=row_num,
                column=2,
                value=teacher.staff_name
            )

            worksheet.cell(
                row=row_num,
                column=3,
                value=teacher.department
            )

            worksheet.cell(
                row=row_num,
                column=4,
                value=levels
            )

            worksheet.cell(
                row=row_num,
                column=5,
                value=str(selected_date)
            )

            worksheet.cell(
                row=row_num,
                column=6,
                value=attendance.status if attendance else "Not Marked"
            )

            worksheet.cell(
                row=row_num,
                column=7,
                value=attendance.remarks if attendance else ""
            )

            worksheet.cell(
                row=row_num,
                column=8,
                value=str(attendance.checked_in_time) if attendance and attendance.checked_in_time else ""
            )

            worksheet.cell(
                row=row_num,
                column=9,
                value=str(attendance.checked_out_time) if attendance and attendance.checked_out_time else ""
            )

            row_num += 1

        # Auto adjust column width
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            worksheet.column_dimensions[column_letter].width = max_length + 5

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response[
            "Content-Disposition"
        ] = f'attachment; filename="Teacher_Attendance_{selected_date}.xlsx"'

        workbook.save(response)

        return response



#  export excel  students attendance api 
class ExportStudentAttendanceExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")
        class_id = request.query_params.get("class_id")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        classes = ClassModel.objects.all()

        if class_id:
            classes = classes.filter(id=class_id)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Student Attendance"

        headers = [
            "Class Code",
            "Class Name",
            "Section",
            "Total Students",
            "Present",
            "Absent",
            "Half Day",
            "Attendance Taken By",
            "Class Incharge"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        row_num = 2

        for cls in classes:

            total_students = StudentAcademicDetails.objects.filter(
                student_class=cls
            ).count()

            attendance_qs = StudentAttendance.objects.filter(
                students_class=cls,
                date=date
            )

            present_count = attendance_qs.filter(
                status="Present"
            ).count()

            absent_count = attendance_qs.filter(
                status="Absent"
            ).count()

            half_day_count = attendance_qs.filter(
                status="Half Day"
            ).count()

            taken_by = attendance_qs.first().taken_by if attendance_qs.exists() else None

            worksheet.cell(
                row=row_num,
                column=1,
                value=getattr(cls, "class_id", "")
            )

            worksheet.cell(
                row=row_num,
                column=2,
                value=getattr(cls, "class_name", "")
            )

            worksheet.cell(
                row=row_num,
                column=3,
                value=getattr(cls, "section", "")
            )

            worksheet.cell(
                row=row_num,
                column=4,
                value=total_students
            )

            worksheet.cell(
                row=row_num,
                column=5,
                value=present_count
            )

            worksheet.cell(
                row=row_num,
                column=6,
                value=absent_count
            )

            worksheet.cell(
                row=row_num,
                column=7,
                value=half_day_count
            )

            worksheet.cell(
                row=row_num,
                column=8,
                value=taken_by.staff_name if taken_by else ""
            )

            worksheet.cell(
                row=row_num,
                column=9,
                value=cls.class_teacher.staff_name
                if hasattr(cls, "class_teacher") and cls.class_teacher
                else ""
            )

            row_num += 1

        # Auto-adjust column widths
        for column in worksheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            worksheet.column_dimensions[column_letter].width = max_length + 5

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        filename = f"Student_Attendance_{date}.xlsx"

        if class_id:
            filename = f"Student_Attendance_Class_{class_id}_{date}.xlsx"

        response[
            "Content-Disposition"
        ] = f'attachment; filename="{filename}"'

        workbook.save(response)

        return response


# staffs attendance export excel api 
class ExportStaffAttendanceExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.query_params.get("date")

        if not date:
            return Response({
                "status": False,
                "message": "Date query parameter is required in YYYY-MM-DD format"
            }, status=status.HTTP_400_BAD_REQUEST)

        staff_members = StaffManagementModel.objects.filter(
            is_teacher=False
        ).select_related("profiles")

        attendance_records = TeacherstaffAttendance.objects.filter(
            date=date,
            teacher__is_teacher=False
        )

        attendance_dict = {
            attendance.teacher_id: attendance
            for attendance in attendance_records
        }

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Staff Attendance"

        headers = [
            "Staff ID",
            "Staff Name",
            "Department",
            "Date",
            "Status",
            "Remarks",
            "Check In Time",
            "Check Out Time"
        ]

        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        row_num = 2

        for staff in staff_members:

            attendance = attendance_dict.get(staff.id)

            worksheet.cell(
                row=row_num,
                column=1,
                value=staff.profiles.user_id if staff.profiles else ""
            )

            worksheet.cell(
                row=row_num,
                column=2,
                value=staff.staff_name
            )

            worksheet.cell(
                row=row_num,
                column=3,
                value=staff.department
            )

            worksheet.cell(
                row=row_num,
                column=4,
                value=str(date)
            )

            worksheet.cell(
                row=row_num,
                column=5,
                value=attendance.status if attendance else "Not Marked"
            )

            worksheet.cell(
                row=row_num,
                column=6,
                value=attendance.remarks if attendance else ""
            )

            worksheet.cell(
                row=row_num,
                column=7,
                value=str(attendance.checked_in_time)
                if attendance and attendance.checked_in_time
                else ""
            )

            worksheet.cell(
                row=row_num,
                column=8,
                value=str(attendance.checked_out_time)
                if attendance and attendance.checked_out_time
                else ""
            )

            row_num += 1

        # Auto-adjust column widths
        for column in worksheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                if cell.value:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[column_letter].width = max_length + 5

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        response[
            "Content-Disposition"
        ] = f'attachment; filename="Staff_Attendance_{date}.xlsx"'

        workbook.save(response)

        return response




#************************************************ dashboard views  ************************************
# dashboard kpi cards 
class DashboardKpiCardsAPIViews(APIView):
    permission_classes = [IsAuthenticated]

    def calculate_growth(self, current, previous):

        # No data in both years
        if current == 0 and previous == 0:
            return 0

        # New records added this year
        if previous == 0 and current > 0:
            return 100

        # Data existed before but now dropped to zero
        if previous > 0 and current == 0:
            return -100

        growth = ((current - previous) / previous) * 100

    # Cap maximum growth at 100%
        if growth > 100:
            growth = 100

        if growth < -100:
            growth = -100

        return round(growth, 2)

    def get(self, request):

        # print("Request:", request)

       
        # print("Authorization Header:", request.headers.get("Authorization"))

     
        # print("All Headers:")
        # for key, value in request.headers.items():
        #     print(f"{key}: {value}")

        # print("Cookies:", request.COOKIES)

        # print("Access Token Cookie:", request.COOKIES.get("access_token"))
        # print("Refresh Token Cookie:", request.COOKIES.get("refresh_token"))

     
        # print("Authenticated User:", request.user)

        # print("request.auth:", request.auth)

        current_year = date.today().year
        previous_year = current_year - 1

        # ==========================
        # TOTAL STUDENTS
        # ==========================

        total_students = Profiles.objects.filter(
            role="Student"
        ).count()

        previous_students = Profiles.objects.filter(
            role="Student",
            date_joined__year__lte=previous_year
        ).count()

        student_growth = self.calculate_growth(
            total_students,
            previous_students
        )

        # ==========================
        # TEACHERS
        # ==========================

        total_teachers = StaffManagementModel.objects.filter(
            is_teacher=True
        ).count()

        previous_teachers = StaffManagementModel.objects.filter(
            is_teacher=True,
            profiles__date_joined__year__lte=previous_year
        ).count()

        teacher_growth = self.calculate_growth(
            total_teachers,
            previous_teachers
        )

        # ==========================
        # NON TEACHING STAFF
        # ==========================

        total_non_teaching = StaffManagementModel.objects.filter(
            is_teacher=False
        ).count()

        previous_non_teaching = StaffManagementModel.objects.filter(
            is_teacher=False,
            profiles__date_joined__year__lte=previous_year
        ).count()

        non_teaching_growth = self.calculate_growth(
            total_non_teaching,
            previous_non_teaching
        )

        # ==========================
        # ACTIVE STUDENTS
        # ==========================

        total_active_students = Profiles.objects.filter(
            role="Student",
            is_active=True
        ).count()

        previous_active_students = Profiles.objects.filter(
            role="Student",
            is_active=True,
            date_joined__year__lte=previous_year
        ).count()

        active_student_growth = self.calculate_growth(
            total_active_students,
            previous_active_students
        )

        teacher_count_change = total_teachers - previous_teachers

        # ==========================
        # NEW ADMISSIONS
        # ==========================

        current_year_admissions = Profiles.objects.filter(
            role="Student",
            date_joined__year=current_year
        ).count()

        previous_year_admissions = Profiles.objects.filter(
            role="Student",
            date_joined__year=previous_year
        ).count()

        admission_growth = self.calculate_growth(
            current_year_admissions,
            previous_year_admissions
        )

        # ==========================
        # BATCHES
        # ==========================

        total_batches = (
            ClassModel.objects
            .exclude(batch__isnull=True)
            .exclude(batch="")
            .values('batch')
            .distinct()
            .count()
        )

        previous_batches = (
            ClassModel.objects
            .filter(created_at__year__lte=previous_year)
            .exclude(batch__isnull=True)
            .exclude(batch="")
            .values('batch')
            .distinct()
            .count()
        )

        batch_growth = self.calculate_growth(
            total_batches,
            previous_batches
        )

        response_data = {
            "total_students": {
                "count": total_students,
                "growth_percentage": student_growth
            },
            "teachers": {
                "count": total_teachers,
                "growth_percentage": teacher_count_change
            },
            "non_teaching_staff": {
                "count": total_non_teaching,
                "growth_percentage": non_teaching_growth
            },
            "active_students": {
                "count": total_active_students,
                "growth_percentage": active_student_growth
            },
            "new_admissions": {
                "count": current_year_admissions,
                "growth_percentage": admission_growth
            },
            "batches": {
                "count": total_batches,
                "growth_percentage": batch_growth
            }
        }

        return Response(
            {
                "status": True,
                "message": "Dashboard data fetched successfully",
                "data": response_data
            },
            status=status.HTTP_200_OK
        )

    # def get(self, request):

    #     user = request.user
       
    #     today = timezone.now().date()

    #     total_students = Profiles.objects.filter(
    #         role="Student").count()
    #     total_teachers = StaffManagementModel.objects.filter(
    #         designation__icontains="Teacher", is_teacher=True).count()
    #     total_non_teaching_staffs = StaffManagementModel.objects.filter(
    #         is_teacher=False).count()
    #     total_active_students = StudentPersonalDetails.objects.filter(
    #         user__is_active=True).count()
    #     new_admission_count = StudentPersonalDetails.objects.filter(user__date_joined = today).count()
    #     total_batches = ClassModel.objects.count()

    #     return Response({
    #         "status": True,
    #         "message": "Dashboard KPI cards data retrieved successfully",
    #         "data": {
    #             "total_students": total_students,
    #             "total_teachers": total_teachers,
    #             "total_non_teaching_staffs": total_non_teaching_staffs,
    #             "total_active_students": total_active_students,
    #             "new_admission_count": new_admission_count,
    #             "total_batches": total_batches
    #         }
    #     }, status=status.HTTP_200_OK)
    


# student admission trend pie chart current years students data for pie chart  based on level wise(high school , lp , up)
class DashboardStudentPieChartBatchCountAPIViews(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        # Pie Chart Data (Class / Level wise)
        class_wise_students = (
            StudentAcademicDetails.objects
            .filter(student_class__isnull=False)
            .values(
                "student_class__level"
            )
            .annotate(
                count=Count("id")
            )
            .order_by()
        )

        pie_chart_data = []

        for item in class_wise_students:
            pie_chart_data.append({
                "label": item["student_class__level"],
                "value": item["count"]
            })

        # Bar Chart Data (Batch Wise)
        batch_wise_students = (
            StudentAcademicDetails.objects
            .exclude(batch__isnull=True)
            .exclude(batch="")
            .values("batch")
            .annotate(
                count=Count("id")
            )
            .order_by("batch")
        )

        bar_chart_data = []

        for item in batch_wise_students:
            bar_chart_data.append({
                "batch": item["batch"],
                "count": item["count"]
            })

        return Response({
            "status": True,
            "message": "Dashboard data fetched successfully",

            "student_admission_trend": pie_chart_data,

            "batch_wise_student_count": bar_chart_data
        }, status=status.HTTP_200_OK)


# line char
class MonthlyCollectionAPIView(APIView):

    def get(self, request):
        current_year = timezone.now().year

        collections = (
            StudentFinancialDetails.objects
            .filter(created_at__year=current_year)
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(total_revenue=Sum('paid_amount'))
            .order_by('month')
        )

        month_data = {
            "Jan": 0,
            "Feb": 0,
            "Mar": 0,
            "Apr": 0,
            "May": 0,
            "Jun": 0,
            "Jul": 0,
            "Aug": 0,
            "Sep": 0,
            "Oct": 0,
            "Nov": 0,
            "Dec": 0,
        }

        for item in collections:
            month_name = item["month"].strftime("%b")
            month_data[month_name] = float(item["total_revenue"] or 0)

        return Response({
            "status": True,
            "message": "Data listed successfully",
            "year": current_year,
            "data": [
                {
                    "month": month,
                    "revenue": revenue
                }
                for month, revenue in month_data.items()
            ]
        })

# =================================== 
# Profile Settings
# ===================================


class ProfileSettingsAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        try:

            serializer = ProfileSettingsSerializer(
                request.user
            )

            return Response({
                "status": True,
                "message": "Profile fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:

            print("PROFILE FETCH ERROR =", str(e))

            return Response({
                "status": False,
                "message": "Unable to fetch profile details."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        

class ProfileSettingsUpdateAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def put(self, request):

        try:

            user = request.user

            serializer = ProfileSettingsUpdateSerializer(
                user,
                data=request.data,
                partial=True
            )

            if not serializer.is_valid():

                return Response({
                    "status": False,
                    "message": "Invalid data",
                    "errors": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)

            with transaction.atomic():

                serializer.save()

            print(
                f"PROFILE UPDATED: {user.user_id}"
            )

            return Response({
                "status": True,
                "message": "Profile updated successfully.",
                "data": ProfileSettingsSerializer(user).data
            }, status=status.HTTP_200_OK)

        except Exception as e:

            print("PROFILE UPDATE ERROR =", str(e))

            return Response({
                "status": False,
                "message": "Unable to update profile.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def patch(self, request):

        return self.put(request)
    
    # change password 
class ChangePasswordAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = ChangePasswordSerializer(
            data=request.data
        )

        if not serializer.is_valid():

            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        current_password = (
            serializer.validated_data[
                "current_password"
            ]
        )

        if not request.user.check_password(
            current_password
        ):

            return Response({
                "status": False,
                "message":
                "Current password is incorrect."
            }, status=status.HTTP_400_BAD_REQUEST)

        request.user.set_password(
            serializer.validated_data[
                "new_password"
            ]
        )

        request.user.save()

        print(
            f"PASSWORD CHANGED: {request.user.user_id}"
        )
        
        return Response({
            "status": True,
            "message":
            "Password changed successfully."
        }, status=status.HTTP_200_OK)
   # logout api  
class LogoutAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):
        print("Authorization Header:", request.headers.get("Authorization"))

        try:

            refresh_token = request.data.get('refresh')

            if refresh_token:
                try:
                    RefreshToken(refresh_token).blacklist()
                except Exception:
                    pass

            user_agent = request.META.get('HTTP_USER_AGENT', '')
            ip_address = get_client_ip(request)
            browser = get_browser(user_agent)
            device_name = get_device(user_agent)
            location = get_location(ip_address)

            latest_login = LoginHistory.objects.filter(
            user=request.user,
            logout_time__isnull=True
            ).order_by('-created_at').first()

            if latest_login and not latest_login.logout_time:
                latest_login.logout_time = timezone.now()
                latest_login.login_status = 'LOGOUT'
                latest_login.save(update_fields=['logout_time', 'login_status'])

            header = request.META.get('HTTP_AUTHORIZATION', '')
            current_token = header.replace('Bearer ', '').strip()

            try:
                from rest_framework_simplejwt.tokens import AccessToken
                current_jti = AccessToken(current_token)['jti']

                UserSession.objects.filter(
                    user=request.user,
                    session_id=current_jti,
                    is_active=True
                ).update(is_active=False)
            
            except Exception :

                UserSession.objects.filter(
                    user=request.user,
                    is_active=True
                ).update(is_active=False)

            return Response({
                'status': True,
                'message': 'Logged out successfully.'
            }, status=status.HTTP_200_OK)

        except Exception as e:

            print('LOGOUT ERROR =', str(e))

            return Response({'status': False,'message': 'Unable to logout.'}, status=500)        
        

        
# =================================== 
# Profile Settings
# ===================================


class ProfileSettingsAPIView(APIView):
 
    permission_classes = [IsAuthenticated]

    def get(self, request):
        print("Request:", request)

       
        print("Authorization Header:", request.headers.get("Authorization"))

     
        # print("All Headers:")
        # for key, value in request.headers.items():
        #     print(f"{key}: {value}")

        # print("Cookies:", request.COOKIES)

        print("Access Token Cookie:", request.COOKIES.get("access_token"))
        print("Refresh Token Cookie:", request.COOKIES.get("refresh_token"))

        try:

            serializer = ProfileSettingsSerializer(
                request.user
            )

            return Response({
                "status": True,
                "message": "Profile fetched successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        except Exception as e:

            print("PROFILE FETCH ERROR =", str(e))

            return Response({
                "status": False,
                "message": "Unable to fetch profile details."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
class ProfileSettingsUpdateAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def put(self, request):

        try:

            user = request.user

            print("REQUEST DATA =", request.data)

            serializer = ProfileSettingsUpdateSerializer(
                user,
                data=request.data,
                partial=True
            )

            if not serializer.is_valid():

                return Response({
                    "status": False,
                    "message": "Invalid data",
                    "errors": serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)

            print("VALIDATED DATA =", serializer.validated_data)

            with transaction.atomic():

                serializer.save()

                print(
                    "SAVED VALUES =",
                    user.gender,
                    user.address
                )

            print(
                f"PROFILE UPDATED: {user.user_id}"
            )

            return Response({
                "status": True,
                "message": "Profile updated successfully.",
                "data": ProfileSettingsSerializer(user).data
            }, status=status.HTTP_200_OK)

        except Exception as e:

            print("PROFILE UPDATE ERROR =", str(e))

            return Response({
                "status": False,
                "message": "Unable to update profile.",
                "error": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    def patch(self, request):

        return self.put(request)

class ChangePasswordAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = ChangePasswordSerializer(
            data=request.data
        )

        if not serializer.is_valid():

            return Response({
                "status": False,
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        current_password = (
            serializer.validated_data[
                "current_password"
            ]
        )

        if not request.user.check_password(
            current_password
        ):

            return Response({
                "status": False,
                "message":
                "Current password is incorrect."
            }, status=status.HTTP_400_BAD_REQUEST)

        request.user.set_password(
            serializer.validated_data[
                "new_password"
            ]
        )

        request.user.save()

        print(
            f"PASSWORD CHANGED: {request.user.user_id}"
        )
        
        return Response({
            "status": True,
            "message":
            "Password changed successfully."
        }, status=status.HTTP_200_OK)
    
    # logout api 


# correct logout api 
class AllLogoutAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            refresh_token = request.data.get('refresh')

            if refresh_token:
                try:
                    RefreshToken(refresh_token).blacklist()
                except Exception:
                    pass

            user_agent = request.META.get('HTTP_USER_AGENT', '')
            ip_address = get_client_ip(request)
            browser = get_browser(user_agent)
            device_name = get_device(user_agent)
            location = get_location(ip_address)

            latest_login = LoginHistory.objects.filter(
            user=request.user,
            logout_time__isnull=True
            ).order_by('-created_at').first()

            if latest_login and not latest_login.logout_time:
                latest_login.logout_time = timezone.now()
                latest_login.login_status = 'LOGOUT'
                latest_login.save(update_fields=['logout_time', 'login_status'])

            header = request.META.get('HTTP_AUTHORIZATION', '')
            current_token = header.replace('Bearer ', '').strip()

            try:
                from rest_framework_simplejwt.tokens import AccessToken
                current_jti = AccessToken(current_token)['jti']

                UserSession.objects.filter(
                    user=request.user,
                    session_id=current_jti,
                    is_active=True
                ).update(is_active=False)
            
            except Exception :

                UserSession.objects.filter(
                    user=request.user,
                    is_active=True
                ).update(is_active=False)

            response =  Response({
                'status': True,
                'message': 'Logged out successfully.'
            }, status=status.HTTP_200_OK)

            response.delete_cookie('access_token')
            response.delete_cookie('refresh_token')

            return response


        except Exception as e:

            print('LOGOUT ERROR =', str(e))

            return Response({
                'status': False,
                'message': 'Unable to logout.'
            }, status=500)

    


# =========================================
# SECURITY SETTINGS
# ========================================= 

class SecurityDashboardAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        try:

            if request.user.role in [
                'SuperAdmin',
                'Administration staff'
            ]:

                data = {
                    "total_logins": LoginHistory.objects.exclude(
                        login_status='FAILED'
                    ).count(),

                    "total_failed_logins": LoginHistory.objects.filter(
                        login_status='FAILED'
                    ).count(),

                    "active_sessions": UserSession.objects.filter(
                        is_active=True,
                        expires_at__gt=timezone.now()
                    ).count(),

                    "password_resets": PasswordResetAudit.objects.count()
                }

            else:

                data = {
                    "total_logins": LoginHistory.objects.filter(
                        user=request.user
                    ).exclude(
                        login_status='FAILED'
                    ).count(),

                    "total_failed_logins": LoginHistory.objects.filter(
                        user=request.user,
                        login_status='FAILED'
                    ).count(),

                    "active_sessions": UserSession.objects.filter(
                        user=request.user,
                        is_active=True,
                        expires_at__gt=timezone.now()
                    ).count(),

                    "password_resets": PasswordResetAudit.objects.filter(
                        reset_by=request.user
                    ).count()
                }

            serializer = SecurityDashboardSerializer(
                instance=data
            )

            return Response({
                "status": True,
                "message":
                "Security dashboard fetched successfully.",
                "data": serializer.data
            })
        
        

        except Exception as e:

            print(
                "SECURITY DASHBOARD ERROR =",
                str(e)
            )

            return Response({
                "status": False,
                "message":
                "Unable to fetch security dashboard."
            }, status=500)


# LOGIN HISTORY 
class LoginHistoryAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        try:
            if request.user.role in ['SuperAdmin', 'Administration staff']:

                queryset = (
                    LoginHistory.objects
                    .select_related('user')
                    .order_by('-created_at')
                )

            else:

                queryset = (
                    LoginHistory.objects
                    .select_related('user')
                    .filter(user=request.user)
                    .order_by('-created_at')
                )

            # ==========================
            # FILTERS
            # ==========================

            name = request.GET.get('name')
            login_date = request.GET.get('login_date')

            if name:
                queryset = queryset.filter(
                    user__fullname__icontains=name
                )

            if login_date:
                queryset = queryset.filter(
                created_at__date=login_date
                )

            # ==========================
            # PAGINATION
            # ==========================

            paginator = ListPagination()

            page = paginator.paginate_queryset(
                queryset,
                request
            )

            serializer = LoginHistorySerializer(
                page,
                many=True
            )

            return paginator.get_paginated_response(
                serializer.data
            )

        except Exception as e:

            print(
                "LOGIN HISTORY ERROR =",
                str(e)
            )

            return Response({
                "status": False,
                "message":
                "Unable to fetch login history."
            }, status=500)
        

class ExportLoginHistoryExcelAPIView(APIView):

    permission_classes = [IsAuthenticated]

    

    def get(self, request):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Login History'

        headers = [
            'User ID',
            'Full Name',
            'Login Time',
            'Logout Time',
            'Status',
            'IP Address',
            'Browser',
            'Device',
            'Location'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)

        if request.user.role in ['SuperAdmin', 'Administration staff']:

            queryset = LoginHistory.objects.all().order_by('-created_at')

        else:

            queryset = LoginHistory.objects.filter(
                user=request.user
            ).order_by('-created_at')

        row_num = 2

        for item in queryset:

            login_time = (
                timezone.localtime(item.login_time).strftime(
                    '%d-%m-%Y %I:%M:%S %p'
                )
                if item.login_time else ''
            )

            logout_time = (
                timezone.localtime(item.logout_time).strftime(
                    '%d-%m-%Y %I:%M:%S %p'
                )
                if item.logout_time else ''
            )

            sheet.cell(row=row_num, column=1).value = (
                item.user.user_id
                if item.user
                else 'Unknown User'
            )

            sheet.cell(row=row_num, column=2).value = (
                item.user.fullname
                if item.user
                else '-'
            )
            sheet.cell(row=row_num, column=3).value = login_time
            sheet.cell(row=row_num, column=4).value = logout_time
            sheet.cell(row=row_num, column=5).value = item.login_status
            sheet.cell(row=row_num, column=6).value = item.ip_address
            sheet.cell(row=row_num, column=7).value = item.browser
            sheet.cell(row=row_num, column=8).value = item.device_name
            sheet.cell(row=row_num, column=9).value = item.location

            row_num += 1

        for column in sheet.columns:

            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )
                except Exception:
                    pass

            sheet.column_dimensions[column_letter].width = max_length + 5

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="login_history.xlsx"'
        )

        workbook.save(response)

        return response
    

class ActiveSessionAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        try:

            expired_sessions = UserSession.objects.filter(
                is_active=True,
                expires_at__lte=timezone.now()
            ).select_related('user')

            for session in expired_sessions:

                session.is_active = False
                session.save(update_fields=['is_active'])

                latest_login = LoginHistory.objects.filter(
                    user=session.user,
                    logout_time__isnull=True
                ).order_by('-created_at').first()

                if latest_login:

                    latest_login.logout_time = session.expires_at
                    latest_login.login_status = 'LOGOUT'

                    latest_login.save(
                        update_fields=[
                            'logout_time',
                            'login_status'
                        ]
                    )

            if request.user.role in [
                'SuperAdmin',
                'Administration staff'
            ]:

                queryset = (
                    UserSession.objects
                    .select_related('user')
                    .filter(
                        is_active=True,
                        expires_at__gt=timezone.now()
                    )
                    .order_by('-created_at')
                )

            else:

                queryset = (
                    UserSession.objects
                    .select_related('user')
                    .filter(
                        user=request.user,
                        is_active=True,
                        expires_at__gt=timezone.now()
                    )
                    .order_by('-created_at')
                )

            serializer = ActiveSessionSerializer(
                queryset,
                many=True
            )

            return Response({
                'status': True,
                'count': len(serializer.data),
                'data': serializer.data
            })

        except Exception as e:

            print(
                'ACTIVE SESSION ERROR =',
                str(e)
            )

            return Response({
                'status': False,
                'message':
                'Unable to fetch active sessions.'
            }, status=500)
        


class SessionSignOutAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        if not request.data.get('session_id'):
            return Response({
                'status': False,
                'message': 'session_id is required.'
            }, status=400)

        session_id = request.data.get('session_id')

        session = UserSession.objects.filter(
            id=session_id,
            is_active=True
        ).select_related('user').first()

        if not session:
            return Response({
                'status': False,
                'message': 'Session not found.'
            }, status=404)

        if request.user.role not in [
            'SuperAdmin',
            'Administration staff'
        ] and session.user != request.user:

            return Response({
                'status': False,
                'message': 'Permission denied.'
            }, status=403)

        session.is_active = False
        session.save(update_fields=['is_active'])

        latest_login = LoginHistory.objects.filter(
            user=session.user,
            logout_time__isnull=True
        ).order_by('-created_at').first()

        if latest_login:
            latest_login.logout_time = timezone.now()
            latest_login.login_status = 'LOGOUT'
            latest_login.save(
                update_fields=[
                    'logout_time',
                    'login_status'
                ]
            )

        return Response({
            'status': True,
            'message': 'Session signed out successfully.',
            'affected_user': session.user.user_id
        })
    

class LogoutAllDevicesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):

        request.user.token_version += 1
        request.user.save(update_fields=['token_version'])

        count = UserSession.objects.filter(
        user=request.user,
        is_active=True
        ).update(
        is_active=False,
        token_version=request.user.token_version
        )

        open_logins = LoginHistory.objects.filter(
            user=request.user,
            logout_time__isnull=True
        )

        for login in open_logins:

            login.logout_time = timezone.now()
            login.login_status = 'LOGOUT'

            login.save(
                update_fields=[
                    'logout_time',
                    'login_status'
                ]
            )

        return Response({
            'status': True,
            'message': 'All devices logged out successfully.',
            'affected_sessions': count
        })


# class ForceLogoutUserAPIView(APIView):

#     permission_classes = [IsAuthenticated]

#     def post(self, request):

#         if request.user.role not in [
#             'SuperAdmin',
#             'Administration staff'
#         ]:

#             return Response({
#                 'status': False,
#                 'message': 'Permission denied.'
#             }, status=403)

#         target_user_id = request.data.get('user_id')

#         target_user = Profiles.objects.filter(
#             user_id=target_user_id
#         ).first()

#         if not target_user:
#             return Response({
#                 'status': False,
#                 'message': 'User not found.'
#             }, status=404)

#         target_user.token_version += 1
#         target_user.save(update_fields=['token_version'])

#         count = UserSession.objects.filter(
#         user=target_user,
#         is_active=True
#         ).update(
#         is_active=False,
#         token_version=target_user.token_version
#         )

#         open_logins = LoginHistory.objects.filter(
#             user=target_user,
#             logout_time__isnull=True
#         )

#         for login in open_logins:

#             login.logout_time = timezone.now()
#             login.login_status = 'LOGOUT'

#             login.save(
#                 update_fields=[
#                     'logout_time',
#                     'login_status'
#                 ]
#             )

#         return Response({
#             'status': True,
#             'message': 'User force logged out successfully.',
#             'affected_sessions': count
#         })
class ForceLogoutUserAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        if request.user.role not in [
            "SuperAdmin",
            "Administration staff",
        ]:
            return Response(
                {
                    "status": False,
                    "message": "Permission denied."
                },
                status=403
            )

        # Increment token version for all users except the current user
        Profiles.objects.exclude(
            id=request.user.id
        ).update(
            token_version=F("token_version") + 1
        )

        # Deactivate all sessions except the current user's
        affected_sessions = UserSession.objects.exclude(
            user=request.user
        ).filter(
            is_active=True
        ).update(
            is_active=False
        )

        # Close login history except the current user's
        logout_time = timezone.now()

        LoginHistory.objects.exclude(
            user=request.user
        ).filter(
            logout_time__isnull=True
        ).update(
            logout_time=logout_time,
            login_status="LOGOUT"
        )

        return Response(
            {
                "status": True,
                "message": "All other users have been force logged out.",
                "affected_sessions": affected_sessions
            },
            status=status.HTTP_200_OK
        )
    

class ResetPasswordAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        try:

            if request.user.role not in ['SuperAdmin', 'Administration staff']:

                return Response({
                    "status": False,
                    "message":
                    "Permission denied."
                }, status=403)

            serializer = (
                PasswordResetSerializer(
                    data=request.data
                )
            )

            if not serializer.is_valid():

                return Response({
                    "status": False,
                    "errors":
                    serializer.errors
                }, status=400)

            user = Profiles.objects.get(
                user_id=
                serializer.validated_data[
                    'user_id'
                ]
            )

            alphabet = (
                string.ascii_letters +
                string.digits +
                "@#$%&!"
            )

            temporary_password = ''.join(
                secrets.choice(alphabet)
                for _ in range(12)
            )

            user.set_password(
                temporary_password
            )

            user.save()

            email_sent = False

            if (
                serializer.validated_data.get(
                    'send_email',
                    True
                )
                and
                user.email
            ):

                try:

                    send_mail(
                        subject="Password Reset",
                        message=(
                            f"Hello {user.fullname or user.user_id},\n\n"
                            f"Your temporary password is:\n\n"
                            f"{temporary_password}\n\n"
                            f"Please log in and change your password immediately."
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[user.email],
                        fail_silently=False,
                    )

                    email_sent = True

                except Exception as mail_error:

                    print(
                        "MAIL ERROR =",
                        str(mail_error)
                    )

            PasswordResetAudit.objects.create(

                target_user=user,

                reset_by=request.user,

                temporary_password_sent=
                email_sent,

                remarks=
                serializer.validated_data.get(
                    'remarks'
                )
            )

            return Response({
                "status": True,
                "message":
                "Password reset successfully.",
                "temporary_password":
                temporary_password
            })

        except Profiles.DoesNotExist:

            return Response({
                "status": False,
                "message":
                "User not found."
            }, status=404)

        except Exception as e:

            print(
                "RESET PASSWORD ERROR =",
                str(e)
            )

            return Response({
                "status": False,
                "message":
                "Unable to reset password."
            }, status=500)        

class UserListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        try:

            if request.user.role not in [
                'SuperAdmin',
                'Administration staff'
            ]:

                return Response({
                    'status': False,
                    'message': 'Permission denied.'
                }, status=403)

            users = Profiles.objects.all().order_by('fullname')

            data = []

            for user in users:

                data.append({
                    'id': user.id,
                    'user_id': user.user_id,
                    'fullname': user.fullname,
                    'role': user.role,
                    'email': user.email,
                    'photo': (
                        request.build_absolute_uri(user.photo.url)
                        if user.photo else None
                    )
                })

            return Response({
                'status': True,
                'count': len(data),
                'data': data
            })

        except Exception as e:

            print(
                'USER LIST ERROR =',
                str(e)
            )

            return Response({
                'status': False,
                'message': 'Unable to fetch users.'
            }, status=500)

