from django.shortcuts import render
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from .serializers import *
from superadmin_app.models import *
from rest_framework.permissions import AllowAny , IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from . pagination import ListPagination
from rest_framework.pagination import PageNumberPagination
from django.db.models import Q , Sum , F , DecimalField , ExpressionWrapper
from datetime import datetime, date, timedelta
from django.utils import timezone
from django.db.models import Count
from re import search
from django.http import HttpResponse
from openpyxl.styles import Font
from openpyxl import Workbook
from teachers_app.serializers import *
# Create your views here.


#****************************************************** student management section ********************************************************************


class StudentDashboardKPIAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        # Date Filter
        date_param = request.GET.get("date")

        if date_param:
            try:
                filter_date = datetime.strptime(
                    date_param,
                    "%Y-%m-%d"
                ).date()
            except ValueError:
                return Response({
                    "status": False,
                    "message": "Date format should be YYYY-MM-DD"
                }, status=400)
        else:
            filter_date = date.today()

        # Class Filter
        class_id = request.GET.get("class_id")

        # Section Filter
        section = request.GET.get("section")

        # Student Academic queryset
        academic_queryset = StudentAcademicDetails.objects.all()

        if class_id:
            academic_queryset = academic_queryset.filter(
                student_class_id=class_id
            )

        if section:
            academic_queryset = academic_queryset.filter(
                section__iexact=section
            )

        # Total Students
        total_students = academic_queryset.count()

        # Student IDs
        student_ids = academic_queryset.values_list(
            "user_id",
            flat=True
        )

        # Attendance
        attendance_queryset = StudentAttendance.objects.filter(
            student_id__in=student_ids,
            date=filter_date
        )

        present_students = attendance_queryset.filter(
            status="Present"
        ).count()

        absent_students = attendance_queryset.filter(
            status="Absent"
        ).count()

        applications_received = academic_queryset.filter(
            admission_date__year=filter_date.year
        ).count()

        fee_pending = StudentFinancialDetails.objects.filter(
            user_id__in=student_ids
        ).filter(
            Q(balance_amount__gt=0) |
            Q(paid_amount__isnull=True) |
            Q(paid_amount=0)
        ).distinct().count()

        return Response({
            "status": True,
            "data": {
                "date": filter_date,
                "class_id": class_id,
                "section": section,
                "total_students": total_students,
                "present_students": present_students,
                "absent_students": absent_students,
                "applications_received": applications_received,
                "fee_pending": fee_pending
            }
        })
    

class AddStudents(APIView):

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):

        serializer = StudentCreateSerializer(
            data=request.data,
            context={"request": request}
        )

        if serializer.is_valid():

            student = serializer.save()

            return Response(
                {
                    "status": True,
                    "message": "Student created successfully",
                    "data": StudentCreateSerializer(
                        student,
                        context={"request": request}
                    ).data
                },
                status=status.HTTP_201_CREATED
            )

        return Response(
            {
                "status": False,
                "errors": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )


class StudentListAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        queryset = StudentAcademicDetails.objects.select_related(
            'user',
            'student_class'
        ).all()

        # -----------------------
        # SEARCH FILTER
        # -----------------------
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(user__fullname__icontains=search) |
                Q(user__user_id__icontains=search)
            )

        # -----------------------
        # CLASS FILTER
        # -----------------------
        class_id = request.GET.get('class_id')
        if class_id:
            queryset = queryset.filter(student_class_id=class_id)

        # -----------------------
        # SECTION FILTER
        section = request.GET.get('section')
        if section:
            queryset = queryset.filter(section__iexact=section)

        # -----------------------
        # SESSION FILTER
        session = request.GET.get('session')
        if session:
            queryset = queryset.filter(batch__iexact=session)

        # -----------------------
        # ATTENDANCE (NO DATE FILTER)
        # Get latest attendance per student
        # -----------------------
        attendance_qs = StudentAttendance.objects.filter(
            student__in=queryset.values_list('user_id', flat=True)
        ).order_by('student_id', '-date')

        latest_attendance = {}
        for att in attendance_qs:
            if att.student_id not in latest_attendance:
                latest_attendance[att.student_id] = att.status

        # -----------------------
        # SORTING
        # -----------------------
        queryset = queryset.order_by('user__fullname')

        # -----------------------
        # PAGINATION
        # -----------------------
        paginator = ListPagination()
        paginated_queryset = paginator.paginate_queryset(queryset, request)

        serializer = StudentListSerializer(
            paginated_queryset,
            many=True,
            context={
                "attendance_map": latest_attendance
            }
        )

        return paginator.get_paginated_response(serializer.data)
    
    
# automatic arrange roll number 
class ArrangeRollNumbersAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        class_id = request.data.get("class_id")
        section = request.data.get("section")

        if not class_id:
            return Response({
                "status": False,
                "message": "class_id is required"
            }, status=400)

        queryset = StudentAcademicDetails.objects.filter(
            student_class_id=class_id
        )

        if section:
            queryset = queryset.filter(
                section__iexact=section
            )

        queryset = queryset.order_by(
            'user__fullname'
        )

        for index, student in enumerate(queryset, start=1):
            student.roll_number = str(index)
            student.save(update_fields=['roll_number'])

        return Response({
            "status": True,
            "message": "Roll numbers arranged successfully",
            "total_students": queryset.count()
        })
    
# ********************* students overviews ************************************       
class StudentoverviewAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, id):

        try:
            personal = StudentPersonalDetails.objects.select_related(
                "user"
            ).get(user__id=id)

        except StudentPersonalDetails.DoesNotExist:
            return Response(
                {
                    "status": False,
                    "message": "Student not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        student = personal.user

        academic = StudentAcademicDetails.objects.filter(
            user=student
        ).first()

        financial = StudentFinancialDetails.objects.filter(
            user=student
        ).first()

        documents = StudentDocumentDetails.objects.filter(
            user=student
        ).first()

        six_months_ago = timezone.now().date() - timedelta(days=180)

        absent_attendance = StudentAttendance.objects.filter(
            student=student,
            status="Absent",
            date__gte=six_months_ago
        ).order_by("-date")

        leave_details = [
            {
                "date": attendance.date,
                "reason": attendance.reason,
                "remarks": attendance.remarks
            }
            for attendance in absent_attendance
        ]

        data = {

            # Header Section
            "personal_details_id": personal.id,
            "profile_id": student.id,
            "student_id": student.user_id,
            "fullname": student.fullname,

            # Student Photo
            "student_photo": (
                request.build_absolute_uri(
                    documents.student_photo.url
                )
                if documents and documents.student_photo
                else None
            ),

            # Personal Information
            "dob": personal.dob,
            "gender": personal.gender,
            "blood_group": personal.blood_group,
            "address": personal.address,

            # Contact Information
            "email": student.email,
            "phone_number": student.phone_number,

            # Parent Information
            "father_name": personal.father_name,
            "father_phone_number": personal.father_phone_number,
            "father_occupation": personal.father_occupation,

            "mother_name": personal.mother_name,
            "mother_phone_number": personal.mother_phone_number,
            "mother_occupation": personal.mother_occupation,

            # Academic Information
            "class_name": (
                academic.student_class.class_name
                if academic and academic.student_class
                else None
            ),
            "section": academic.section if academic else None,
            "batch": academic.batch if academic else None,
            "roll_number": academic.roll_number if academic else None,
            "student_type": academic.student_type if academic else None,
            "admission_id": academic.admission_id if academic else None,
            "admission_date": academic.admission_date if academic else None,
            "previous_institute": (
                academic.previous_institute
                if academic else None
            ),
            "previous_qualifications": (
                academic.previous_qualifications
                if academic else None
            ),
            "academic_status": (
                academic.status
                if academic else None
            ),

            # Financial Information
            "admission_amount": (
                financial.admission_amount
                if financial else None
            ),
            "course_fee": (
                financial.course_fee
                if financial else None
            ),
            "discount": (
                financial.discount
                if financial else None
            ),
            "paid_amount": (
                financial.paid_amount
                if financial else None
            ),
            "balance_amount": (
                financial.balance_amount
                if financial else None
            ),
            "payment_mode": (
                financial.payment_mode
                if financial else None
            ),
            "installment_plan": (
                financial.installment_plan
                if financial else None
            ),

            # Documents
            "birth_certificate": (
                request.build_absolute_uri(
                    documents.birth_certificate.url
                )
                if documents and documents.birth_certificate
                else None
            ),

            "previous_school_tc": (
                request.build_absolute_uri(
                    documents.previous_school_tc.url
                )
                if documents and documents.previous_school_tc
                else None
            ),

            "aadhar_card": (
                request.build_absolute_uri(
                    documents.aadhar_card.url
                )
                if documents and documents.aadhar_card
                else None
            ),

            "parent_id_proof": (
                request.build_absolute_uri(
                    documents.parent_id_proof.url
                )
                if documents and documents.parent_id_proof
                else None
            ),

            "caste_certificate": (
                request.build_absolute_uri(
                    documents.caste_certificate.url
                )
                if documents and documents.caste_certificate
                else None
            ),

            # Leave Details
            "total_leave_days": absent_attendance.count(),
            "leave_details": leave_details
        }

        return Response(
            {
                "status": True,
                "message": "Student overview fetched successfully",
                "data": data
            },
            status=status.HTTP_200_OK
        )
    
class StudentCheckAdmissionIdAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        admission_id = request.query_params.get("admission_id")

        if not admission_id:
            return Response(
                {
                    "status": False,
                    "message": "Admission ID is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        exists = StudentAcademicDetails.objects.filter(
            admission_id__iexact=admission_id.strip()
        ).exists()

        if exists:
            return Response(
                {
                    "status": False,
                    "exists": True,
                    "message": "Admission ID already exists"
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {
                "status": True,
                "exists": False,
                "message": "Admission ID is available"
            },
            status=status.HTTP_200_OK
        )        
class SectionsdropAPIView(APIView):

    def get(self, request):
        class_id = request.query_params.get("class_id")

        if not class_id:
            return Response({
                "status": False,
                "message": "class_id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            class_obj = ClassModel.objects.get(id=class_id)
        except ClassModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Class not found"
            }, status=status.HTTP_404_NOT_FOUND)

        sections = (
            ClassModel.objects
            .filter(class_name=class_obj.class_name)
            .exclude(section__isnull=True)
            .exclude(section="")
            .values_list("section", flat=True)
            .distinct()
        )

        return Response({
            "status": True,
            "class_id": class_obj.id,
            "class_name": class_obj.class_name,
            "sections": list(sections)
        }, status=status.HTTP_200_OK)


class StudentEditAPIView(APIView):

    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def put(self, request, profile_id):
        print(request.data)

        try:
            student = Profiles.objects.get(
                id=profile_id,
                role="Student"
            )

        except Profiles.DoesNotExist:
            return Response(
                {
                    "status": False,
                    "message": "Student not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        serializer = StudentCreateSerializer(
            student,
            data=request.data,
            partial=True,
            context={
                "request": request
            }
        )

        if serializer.is_valid():

            student = serializer.save()

            return Response(
                {
                    "status": True,
                    "message": "Student updated successfully",
                    "data": StudentCreateSerializer(
                        student
                    ).data
                },
                status=status.HTTP_200_OK
            )

        return Response(
            {
                "status": False,
                "errors": serializer.errors
            },
            status=status.HTTP_400_BAD_REQUEST
        )
    
# classes list for dropdowns 
class ListClassesForDropDowns(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        classes = ClassModel.objects.filter(category=request.user.category).order_by('class_name')

        # seen = set()
        data = []

        for cls in classes:
            if cls.class_name :
                # seen.add(cls.class_name)

                data.append({
                    "id" : cls.id,
                    "class_id": cls.class_id,
                    "class_name": cls.class_name,
                    "class_section": cls.section,
                    "class_batch" : cls.batch,
            
                    # "cla"
                })

        return Response({
            "status": True,
            "message": "Data fetched successfully",
            "data": data
        })



#*************************** Staff management section views ******************************************#####

# staff management kpi cards view 
class StaffManagementKPICards(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        today = timezone.now().date()
        category = request.user.category

        # Base queryset filtered by user's category for data isolation
        staff_qs = StaffManagementModel.objects.filter(profiles__category=category)

        if request.user.role == "SuperAdmin":
            staff_qs = StaffManagementModel.objects.all()


        # Use .count() to return numeric values for KPI cards
        total_staff_count = staff_qs.count()
        teaching_staff_count = staff_qs.filter(is_teacher=True).count()
        non_teaching_staff_count = staff_qs.filter(is_teacher=False).count()

        # Filter attendance by current date, category, and teacher status
        attendance_qs = TeacherstaffAttendance.objects.filter(
            date=today,
            teacher__profiles__category=category,
            teacher__is_teacher=True
        )

        present_teachers = attendance_qs.filter(status='Present').count()
        absent_teachers = attendance_qs.filter(status='Absent').count()

        return Response({
            "status": True,
            "message": "Staff management KPI cards data retrieved successfully",
            "data": {
                "total_staffs": total_staff_count,
                "teaching_staffs": teaching_staff_count,
                "non_teaching_staffs": non_teaching_staff_count,
                "todays_present_teachers": present_teachers,
                "todays_absent_teachers": absent_teachers
            }
        }, status=status.HTTP_200_OK)


class AddStaffManagementView(APIView):

    permission_classes = [IsAuthenticated]
    # parser_classes = [MultiPartParser, FormParser]



    def post(self, request):

        if request.user.role == "Non-administration staff":
            return Response({
                "status": False,
                "message": "You do not have permission to create staff members."
            }, status=status.HTTP_403_FORBIDDEN)

        serializer = StaffCreateSerializer(data=request.data ,  context={'request': request})

        if serializer.is_valid():

            serializer.save()

            return Response({
                "status": True,
                "message": "Staff created successfully",
                "data": serializer.data
            }, status=status.HTTP_201_CREATED)

        return Response({
            "status": False,
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    


# list teaching staff 
class ListTeachingStaffAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        department = request.query_params.get("department")
        if request.user.role == "SuperAdmin":
            staff_members = StaffManagementModel.objects.all()
        else:


            staff_members = StaffManagementModel.objects.filter(
                is_teacher=True,
                profiles__category=request.user.category
            )
        # Optional department filter
        if department:
            staff_members = staff_members.filter(
                department__iexact=department
            )

        paginator = ListPagination()

        paginated_queryset = paginator.paginate_queryset(
                staff_members,
                request
            )

        serializer = ListStaffSerializer(
                paginated_queryset,
                many=True
            )

        return paginator.get_paginated_response(
                serializer.data
            )

# list non teaching staff
class ListNonTeachingStaffAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        if request.user.role == "SuperAdmin":
            staff_members = StaffManagementModel.objects.all()
        else:


            staff_members = StaffManagementModel.objects.filter(is_teacher=False, profiles__category=request.user.category)
        paginator = ListPagination()

        paginated_queryset = paginator.paginate_queryset(
            staff_members,
                request
            )

        serializer = ListStaffSerializer(paginated_queryset, many=True)

        return paginator.get_paginated_response(serializer.data)


# block unblock staffs
class StaffBlockStatusAPIView(APIView):

    def patch(self, request, staff_id):
        try:
            staff = StaffManagementModel.objects.get(id=staff_id)
        except StaffManagementModel.DoesNotExist:
            return Response(
                {
                    "status": False,
                    "message": "Staff not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        is_block = request.data.get("is_block")

        if is_block is None:
            return Response(
                {
                    "status": False,
                    "message": "is_block field is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        staff.is_block = is_block
        staff.save()

        return Response(
            {
                "status": True,
                "message": "Block status updated successfully",
                "data": {
                    "staff_id": staff.id,
                    "staff_name": staff.staff_name,
                    "is_block": staff.is_block
                }
            },
            status=status.HTTP_200_OK
        )


# edit staffs details 
class EditStaffAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def patch(self, request, staff_id):

        try:
            staff_member = StaffManagementModel.objects.get(id=staff_id)
        except StaffManagementModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Staff member not found"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = StaffCreateSerializer(staff_member, data=request.data, partial=True)

        if serializer.is_valid():

            serializer.save()

            return Response({
                "status": True,
                "message": "Staff member updated successfully",
                "data": serializer.data
             }, status=status.HTTP_200_OK)

        return Response({
            "status": False,
            "errors": serializer.errors    
        })
    


#kpi cards api 
# class SatffManagementKPICardsAPIView(APIView):
#     permission_classes = [IsAuthenticated]
#     def get(self, request):

#         total_staffs = StaffManagementModel.objects.filter(profiles__category=request.user.category).count()

#         total_teaching_staff = StaffManagementModel.objects.filter(
#             is_teacher=True,
#             profiles__category=request.user.category
#         ).count()

#         total_non_teaching_staff = StaffManagementModel.objects.filter(
#             is_teacher=False,
#             profiles__category=request.user.category
#         ).count()

#         present_teachers_today = TeacherLeave.objects.filter(
#             teacher__profiles__category=request.user.category,
#             status="Approved",
#             from_date__lte=timezone.now().date(),
#             to_date__gte=timezone.now().date()
#         ).count()

#         absent_teachers_today = TeacherLeave.objects.filter(
#             teacher__profiles__category=request.user.category,
#             status="Pending",
#             from_date__lte=timezone.now().date(),
#             to_date__gte=timezone.now().date()
#         ).count()



#         return Response({
#             "status": True,
#             "message": "Staff management KPI cards fetched successfully",
#             "data": {
#                 "total_staffs": total_staffs,
#                 "total_teaching_staff": total_teaching_staff,
#                 "total_non_teaching_staff": total_non_teaching_staff,
#                 "present_teachers_today": present_teachers_today
#             }
#         }, status=status.HTTP_200_OK)




#**********************Academic management views  *******************************

# kpi cards 
class AcademicManagementKPICardsAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self , request):
        total_students = StudentPersonalDetails.objects.filter(user__category=request.user.category).count()
        # total_teachers = StaffManagementModel.objects.filterrequest.user.category)
        total_teachers = StaffManagementModel.objects.filter(is_teacher=True,profiles__category=request.user.category).count()
        total_administration_staff = StaffManagementModel.objects.filter(profiles__role="Administration staff").count()
        total_non_administration_staff = StaffManagementModel.objects.filter(profiles__role="Non-administration staff").count()


        return Response({
            "status ": True , 
            "Message" : "KPI cards details listed successfully",
            "data" : {
                "total_students": total_students,
                "total_teachers": total_teachers,
                "total_administration_staff": total_administration_staff,
                "total_non_administration_staff": total_non_administration_staff
            }

        })




# class add list api 
class AddListClassAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        # Base queryset with total student count
        classes = ClassModel.objects.annotate(
            total_students=Count("studentacademicdetails")
        )

        # Filter category for non-superadmin users
        if request.user.role != "SuperAdmin":
            classes = classes.filter(
                category=request.user.category
            )

        paginator = ListPagination()

        paginated_classes = paginator.paginate_queryset(
            classes,
            request
        )

        data = []

        for cls in paginated_classes:

            data.append({
                "id": cls.id,
                "class_id": cls.class_id,
                "class_name": cls.class_name,
                "class_batch": cls.batch,
                "level": cls.level,
                "section": cls.section,
                "total_students": cls.total_students,
                "status": cls.status,
                "department": cls.department,
                "branch": cls.branch,
            })

        return paginator.get_paginated_response(data)

    def post(self, request):

        serializer = AddListEditClassSerializer(data=request.data)

        if serializer.is_valid():

            serializer.save()
            category = request.user.category
            class_instance = serializer.instance
            class_instance.category = category
            class_instance.save()


            return Response({
                "status": True,
                "message": "Class created successfully",
                "data": serializer.data
             }, status=status.HTTP_201_CREATED)

        return Response({
            "status": False,
            "errors": serializer.errors    
        },status=status.HTTP_400_BAD_REQUEST)


# edit class  api 

class EditClassAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def patch(self, request, class_id):

        try:
            class_instance = ClassModel.objects.get(id=class_id, category=request.user.category)
        except ClassModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Class not found"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = AddListEditClassSerializer(class_instance, data=request.data, partial=True)

        if serializer.is_valid():

            serializer.save()

            return Response({
                "status": True,
                "message": "Class updated successfully",
                "data": serializer.data
             }, status=status.HTTP_200_OK)

        return Response({
            "status": False,
            "errors": serializer.errors    
        })


 # class details by class id 
class ClassDetailsById(APIView):
    permission_classes = [IsAuthenticated]
    def get(self, request, class_id):

        try:
            class_id = ClassModel.objects.get(id=class_id, category=request.user.category)
        except ClassModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Class not found"
            }, status=status.HTTP_404_NOT_FOUND)
        serializers =   AddListEditClassSerializer(class_id)
        return Response({
            "status": True,
            "message": "Class details fetched successfully",
            "data": serializers.data})
        
        

# filter classes 

class FilterClassByBatchAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, batch):

        classes = ClassModel.objects.filter(batch=batch, category=request.user.category)
        paginator = ListPagination()

        paginated_classes = paginator.paginate_queryset(
            classes,
            request
        )

        data = []

        for class_instance in paginated_classes:

            data.append({
                "id": class_instance.id,
                "class_id": class_instance.class_id,
                "class_name": class_instance.class_name,
                "total_students": StudentAcademicDetails.objects.filter(
        student_class=class_instance
    ).count(),

                "subject": list(
                    class_instance.subject.values_list(
                        "name",
                        flat=True
                    )
                ),
                "level": class_instance.level,
                "section": class_instance.section,
                "class_teacher": class_instance.class_teacher.staff_name  if class_instance.class_teacher else None,
                "status": class_instance.status,
                "batch": class_instance.batch,
                "department": class_instance.department,
                "branch": class_instance.branch,
                "created_at": class_instance.created_at,
                "updated_at": class_instance.updated_at
            })

        return paginator.get_paginated_response(data)
    

# list sections by class name 
class ClassSectionsAPIView(APIView):

    def get(self, request):
        class_id = request.query_params.get("class_id")

        if not class_id:
            return Response({
                "status": False,
                "message": "class_id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            class_obj = ClassModel.objects.get(id=class_id)
        except ClassModel.DoesNotExist:
            return Response({
                "status": False,
                "message": "Class not found"
            }, status=status.HTTP_404_NOT_FOUND)

        sections = (
            ClassModel.objects
            .filter(class_name=class_obj.class_name)
            .exclude(section__isnull=True)
            .exclude(section="")
            .values_list("section", flat=True)
            .distinct()
        )

        return Response({
            "status": True,
            "class_id": class_obj.id,
            "class_name": class_obj.class_name,
            "sections": list(sections)
        }, status=status.HTTP_200_OK)
    

# add time table for teachers 
class AddTeacherTimeTableAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        serializer = TeacherTimeTableSerializer(data=request.data)

        if serializer.is_valid():

            serializer.save()
            category = request.user.category
            class_instance = serializer.instance
            class_instance.category = category
            class_instance.save()

            return Response({
                "status": True,
                "message": "Time table created successfully",
                "data": serializer.data
             }, status=status.HTTP_201_CREATED)

        return Response({
            "status": False,
            "errors": serializer.errors    
        })
    

# list each teachers timetable 
class ListTeacherTimeTableAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request , teacher_id):

        if not teacher_id:
            return Response({
                "status": False,
                "message": "teacher_id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        timetables = TeachersTimeTable.objects.filter(
            teacher_id=teacher_id,
            teacher__profiles__category=request.user.category
        ).order_by("day_of_week", "period")

        if not timetables.exists():
            return Response({
                "status": False,
                "message": "No timetable found for this teacher"
            }, status=status.HTTP_404_NOT_FOUND)

        timetable_data = {
            "Monday": [],
            "Tuesday": [],
            "Wednesday": [],
            "Thursday": [],
            "Friday": [],
            "Saturday": [],
            "Sunday": []
        }

        teacher_name = timetables.first().teacher.staff_name

        for timetable in timetables:

            timetable_data[timetable.day_of_week].append({
                "id": timetable.id,
                "period": timetable.period,
                "subject": timetable.subject,
                "class_assigned": (
                    timetable.class_assigned.class_name
                    if timetable.class_assigned
                    else None
                ),
                "class_id" : timetable.class_assigned.id if timetable.class_assigned else None,
                "class_section": timetable.class_assigned.section if timetable.class_assigned else None,
                "class_batch" : timetable.class_assigned.batch if timetable.class_assigned else None,
                
            })

        return Response({
            "status": True,
            "message": "Timetable fetched successfully",
            "teacher": teacher_name,
            "timetable": timetable_data
        }, status=status.HTTP_200_OK)
    


# edit teachers timetable 
class EditTeacherTimeTableAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def patch(self, request, timetable_id):

        try:
            timetable_instance = TeachersTimeTable.objects.get(id=timetable_id, teacher__profiles__category=request.user.category)
        except TeachersTimeTable.DoesNotExist:
            return Response({
                "status": False,
                "message": "Timetable entry not found"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = TeacherTimeTableSerializer(timetable_instance, data=request.data, partial=True)

        if serializer.is_valid():

            serializer.save()

            return Response({
                "status": True,
                "message": "Timetable entry updated successfully",
                "data": serializer.data
             }, status=status.HTTP_200_OK)

        return Response({
            "status": False,
            "errors": serializer.errors    
        })


#delete teachers time table 
class DeleteTeachersTimeTable(APIView):
    permission_classes = [IsAuthenticated]

    def delete(self, request, time_table_id):

        if not time_table_id:
            return Response(
                {
                    "status": False,
                    "message": "time_table_id is required"
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            timetable = TeachersTimeTable.objects.get(id=time_table_id)

        except TeachersTimeTable.DoesNotExist:
            return Response(
                {
                    "status": False,
                    "message": "Time table not found"
                },
                status=status.HTTP_404_NOT_FOUND
            )

        timetable.delete()

        return Response(
            {
                "status": True,
                "message": "Time table deleted successfully"
            },
            status=status.HTTP_200_OK
        )




# list department for dropdown for filtering teachers   
class DepartmentListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        departments = (

            StaffManagementModel.objects

            .exclude(department__isnull=True)

            .exclude(department__exact='')

            .values_list('department', flat=True)

            .distinct()

            .order_by('department')

        )

        return Response({

            "status": True,

            "count": len(departments),

            "data": [

                {

                    "department": department

                }

                for department in departments

            ]

        })

#*************************** Assign substitute teacher section views ******************************************#####
# todays absent teachers list api 

class TodaysAbsentTeachersAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):
         
        today = timezone.now().date()
        absent_teachers_qs = TeacherLeave.objects.filter(teacher__profiles__category=request.user.category,
            from_date__lte=today,
            to_date__gte=today,
            # status="Approved"
        ).select_related(
            "teacher"
        )

        if request.user.role == "SuperAdmin":
            absent_teachers_qs = TeacherLeave.objects.filter(
                from_date__lte=today,
                to_date__gte=today,
                # status="Approved"
            ).select_related(
                "teacher"
            )


        absent_teachers = []

        for leave in absent_teachers_qs:

            teacher = leave.teacher

            absent_teachers.append({
                "teacher_id": teacher.id,
                "role": teacher.profiles.role,
                "teacher_name": teacher.staff_name,
                "user_id": teacher.profiles.user_id,
                "leave_start_date": leave.from_date,
                "leave_end_date": leave.to_date,
                "leave_reason": leave.reason
            })

        return Response({
            "status": True,
            "message": "Today's absent teachers fetched successfully",
            "category": request.user.category.name if request.user.category else None,
            "date": today,
            "absent_teachers_count": len(absent_teachers),
            "absent_teachers": absent_teachers,
         
        }, status=status.HTTP_200_OK)  



  # get teacher todays timetable by teacher id 
class TeacherTodaysTimeTableAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, teacher_id):

        # Get current day of the week (e.g., Monday, Tuesday)
        today = timezone.now().strftime('%A')

        timetables = TeachersTimeTable.objects.filter(
            teacher=teacher_id,
            day_of_week=today,
            teacher__profiles__category=request.user.category
        ).select_related(
            "teacher",
            "class_assigned"
        ).order_by('period')

        if not timetables.exists():
            return Response({
                "status": False,
                "message": f"No timetable found for this teacher on {today}"
            }, status=status.HTTP_404_NOT_FOUND)
        
        

        timetable_data = []
        for item in timetables:
            timetable_data.append({
                "id": item.id,
                "day_of_week": item.day_of_week,
                "period": item.period,
                "subject": item.subject,
                "teacher_name": item.teacher.staff_name,
                "class_assigned": {
                    "id": item.class_assigned.id,
                    "class_name": item.class_assigned.class_name,
                    "section": item.class_assigned.section
                }
            })

        return Response({
            "status": True,
            "message": f"Today's ({today}) timetable fetched successfully",
            "data": timetable_data
        }, status=status.HTTP_200_OK)



# list the avilable teachers for substitute teacher assignment api
class AvailableSubstituteTeachersAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        teacher_id = request.data.get("teacher_id")
        period = request.data.get("period")
        class_id = request.data.get("class_id")

        if not teacher_id or not period or not class_id:
            return Response({
                "status": False,
                "message": "teacher_id, period and class_id are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        today = timezone.now().date()
        today_day = today.strftime("%A")

        # Teachers on leave today
        absent_teacher_ids = TeacherLeave.objects.filter(
            from_date__lte=today,
            to_date__gte=today,
            status="Approved"
        ).values_list("teacher_id", flat=True)

        # Teachers already handling a class in this period
        busy_teacher_ids = TeachersTimeTable.objects.filter(
            day_of_week=today_day,
            period=period
        ).values_list("teacher_id", flat=True)

        available_teachers = StaffManagementModel.objects.filter(
            profiles__category=request.user.category,profiles__role="Administration staff",is_teacher=True
        ).exclude(
            id__in=busy_teacher_ids
        ).exclude(
            id__in=absent_teacher_ids
        ).exclude(
            id=teacher_id
        ).select_related("profiles").order_by("staff_name")

        data = []

        for teacher in available_teachers:
            data.append({
                "teacher_id": teacher.id,
                "teacher_name": teacher.staff_name,
                "user_id": teacher.profiles.user_id
            })

        return Response({
            "status": True,
            "message": "Available substitute teachers fetched successfully",
            "teacher_id": teacher_id,
            "period": period,
            "class_id": class_id,
            "available_teachers_count": len(data),
            "available_teachers": data
        })


# assign substitute teacher api
class AssignSubstituteTeacherAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def post(self, request):

        class_id = request.data.get("class_id")
        date = request.data.get("date")
        period = request.data.get("period")
        original_teacher_id = request.data.get("original_teacher_id")
        substitute_teacher_id = request.data.get("substitute_teacher_id")

        if not all([
            class_id,
            date,
            period,
            original_teacher_id,
            substitute_teacher_id
        ]):
            return Response({
                "status": False,
                "message": "class_id, date, period, original_teacher_id and substitute_teacher_id are required"
            }, status=status.HTTP_400_BAD_REQUEST)

        assigned_class = ClassModel.objects.filter(
            id=class_id
        ).first()

        if not assigned_class:
            return Response({
                "status": False,
                "message": "Class not found"
            }, status=status.HTTP_404_NOT_FOUND)

        original_teacher = StaffManagementModel.objects.filter(
            id=original_teacher_id
        ).first()

        if not original_teacher:
            return Response({
                "status": False,
                "message": "Original teacher not found"
            }, status=status.HTTP_404_NOT_FOUND)

        substitute_teacher = StaffManagementModel.objects.filter(
            id=substitute_teacher_id
        ).first()

        if not substitute_teacher:
            return Response({
                "status": False,
                "message": "Substitute teacher not found"
            }, status=status.HTTP_404_NOT_FOUND)

        timetable = TeachersTimeTable.objects.filter(
            teacher=original_teacher,
            class_assigned=assigned_class,
            period=period
        ).first()

        subject = timetable.subject if timetable else ""

        assignment, created = SubstituteTeacherAssignment.objects.update_or_create(
            class_assigned=assigned_class,
            date=date,
            period=period,
            defaults={
                "subject": subject,
                "original_teacher": original_teacher,
                "substitute_teacher": substitute_teacher
            }
        )

        # Notification for substitute teacher
        Notification.objects.create(
            user=substitute_teacher.profiles,
            title="Substitute Class Assigned",
            message=(
                f"You have been assigned to take "
                f"{assigned_class.class_name} ({assigned_class.section}) "
                f"on {date}, Period {period}."
                f"{f' Subject: {subject}.' if subject else ''}"
            )
        )

        # Notification for original teacher
        Notification.objects.create(
            user=original_teacher.profiles,
            title="Substitute Teacher Assigned",
            message=(
                f"{substitute_teacher.staff_name} has been assigned as your "
                f"substitute teacher for {assigned_class.class_name} "
                f"({assigned_class.section}) on {date}, Period {period}."
                f"{f' Subject: {subject}.' if subject else ''}"
            )
        )

        return Response({
            "status": True,
            "message": "Substitute teacher assigned successfully",
            "assignment_id": assignment.id,
            "created": created
        })
    

# list the subtitution teacher assignment for a particular date api
class ListSubstituteTeacherAssignmentAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):

        date = request.GET.get("date")

        if not date:
            return Response({
                "status": False,
                "message": "date is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        assignments = SubstituteTeacherAssignment.objects.filter(
            date=date,
            class_assigned__category=request.user.category
        ).select_related(
            "class_assigned",
            "original_teacher",
            "substitute_teacher"
        ).order_by("period")

        data = []

        for assignment in assignments:
            data.append({
                "assignment_id": assignment.id,
                "class_id": assignment.class_assigned.id,
                "class_name": assignment.class_assigned.class_name,
                "period": assignment.period,
                "subject": assignment.subject,
                "original_teacher_id": assignment.original_teacher.id,
                "original_teacher_name": assignment.original_teacher.staff_name,
                "substitute_teacher_id": assignment.substitute_teacher.id,
                "substitute_teacher_name": assignment.substitute_teacher.staff_name,
                "reason": assignment.reason
            })

        return Response({
            "status": True,
            "message": f"Substitute teacher assignments for {date} fetched successfully",
            "assignments_count": len(data),
            "assignments": data
        }, status=status.HTTP_200_OK)
    

# get subsititude details by substitute assignment id api 
class SubstituteTeacherAssignmentDetailAPIView(APIView):

    permission_classes = [IsAuthenticated]
    def get(self, request, assignment_id):

        try:
            assignment = SubstituteTeacherAssignment.objects.select_related(
                "class_assigned",
                "original_teacher",
                "substitute_teacher"
            ).get(id=assignment_id, class_assigned__category=request.user.category)

            data = {
                "assignment_id": assignment.id,
                "class_id": assignment.class_assigned.id,
                "class_name": assignment.class_assigned.class_name,
                "period": assignment.period,
                "subject": assignment.subject,
                "original_teacher_id": assignment.original_teacher.id,
                "original_teacher_name": assignment.original_teacher.staff_name,
                "substitute_teacher_id": assignment.substitute_teacher.id,
                "substitute_teacher_name": assignment.substitute_teacher.staff_name,
                "reason": assignment.reason
            }

            return Response({
                "status": True,
                "message": "Substitute teacher assignment details fetched successfully",
                "data": data
            }, status=status.HTTP_200_OK)

        except SubstituteTeacherAssignment.DoesNotExist:
            return Response({
                "status": False,
                "message": "Substitute teacher assignment not found"
            }, status=status.HTTP_404_NOT_FOUND)
        
# list all teachers leave api 
class ListAllTeachersLeaveRequests(APIView):
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
           
           teachers = TeacherLeave.objects.all()
           serializer = TeachersLeaveListSerializer(teachers, many=True)
           return Response({
               
               "status": True ,
                "message": "Teachers leave requests fetched successfully", 
                "data": serializer.data        })



# get details of leave \
class LeaveDetailsAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        leave_id = request.query_params.get("leave_id")

        if not leave_id:
            return Response({
                "status": False,
                "message": "leave_id is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            leave = TeacherLeave.objects.get(id=leave_id)
        except TeacherLeave.DoesNotExist:
            return Response({
                "status": False,
                "message": "Leave not found"
            }, status=status.HTTP_404_NOT_FOUND)

        serializer = TeachersLeaveListSerializer(leave)
        return Response({
            "status": True,
            "message": "Leave details fetched successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)
    



# approve  and reject leave
class ApproveRejectTeachersLeaveRequests(APIView):
    permission_classes = [IsAuthenticated]

    def post(self , request):

        leave_id = request.data.get("leave_id")
        action = request.data.get("action")

        if not leave_id or not action:
            return Response({
                "status": False,
                "message": "leave_id and action are required"
            }, status=status.HTTP_400_BAD_REQUEST)
        leave_id = TeacherLeave.objects.get(id=leave_id)

        if action == "approve":
            leave_id.status = "Approved"
        elif action == "reject":
            leave_id.status = "Rejected"
        else:
            return Response({
                "status": False,
                "message": "Invalid action" 
                }, status=status.HTTP_400_BAD_REQUEST)
        leave_id.save()

        return Response({
            "status": True,
            "message": "Teachers leave request updated successfully"
            },
            status=status.HTTP_200_OK)



           

# ==================================================================================================
# FINANCE MANAGEMENT APIs
# =================================================================================================

class FinanceDashboardAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        total_admission = StudentAcademicDetails.objects.count()

        total_amount_collected = StudentFinancialDetails.objects.aggregate(
            total=Sum('paid_amount')
        )['total'] or 0

        pending_amount = StudentFinancialDetails.objects.aggregate(
            total=Sum('balance_amount')
        )['total'] or 0

        return Response({
            'status': True,
            'data': {
                'total_admission': total_admission,
                'total_amount_collected': total_amount_collected,
                'pending_amount': pending_amount
            }
        }, status=status.HTTP_200_OK)


class AdmissionListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        queryset = StudentFinancialDetails.objects.select_related(
            'user'
        ).order_by('-id')
        
        queryset = queryset.filter(user__isnull=False)

        search = request.GET.get('search')
        admission_id = request.GET.get('admission_id')
        admission_date = request.GET.get('admission_date')

        if admission_id:
            queryset = queryset.filter(
                user__studentacademicdetails__admission_id__icontains=admission_id
            )

        if admission_date:
            queryset = queryset.filter(
                user__studentacademicdetails__admission_date=admission_date
            )

        if search:
            queryset = queryset.filter(
                Q(user__fullname__icontains=search) |
                Q(user__studentacademicdetails__admission_id__icontains=search)
            )

        paginator = PageNumberPagination()
        paginator.page_size = int(request.GET.get('page_size', 10))

        result_page = paginator.paginate_queryset(queryset, request)

        serializer = AdmissionListSerializer(result_page, many=True)

        return paginator.get_paginated_response({
            'status': True,
            'data': serializer.data
        })


class AdmissionDetailAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request, pk):

        try:
            student = Profiles.objects.select_related(
                'studentpersonaldetails',
                'studentacademicdetails',
                'studentfinancialdetails'
            ).get(pk=pk)

            print(
                "DETAIL STUDENT =",
                student.id,
                student.fullname
            )

            serializer = AdmissionDetailSerializer(student)

            return Response({
                'status': True,
                'data': serializer.data
            }, status=status.HTTP_200_OK)

        except Profiles.DoesNotExist:

            return Response({
                'status': False,
                'message': 'Student not found'
            }, status=status.HTTP_404_NOT_FOUND)

 
# admission update 
class AdmissionUpdateAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get_student(self, pk):
        try:
            return Profiles.objects.select_related(
                'studentpersonaldetails',
                'studentacademicdetails',
                'studentfinancialdetails'
            ).get(pk=pk)

        except Profiles.DoesNotExist:
            return None

    def put(self, request, pk):

        student = self.get_student(pk)

        if not student:
            return Response({
                'status': False,
                'message': 'Student not found'
            }, status=status.HTTP_404_NOT_FOUND)

        print("UPDATE REQUEST DATA =", request.data)

        serializer = AdmissionUpdateSerializer(
            student,
            data=request.data,
            partial=True
        )

        if serializer.is_valid():

            serializer.save()

            student.refresh_from_db()

            detail_serializer = AdmissionDetailSerializer(student)

            return Response({
                'status': True,
                'message': 'Admission updated successfully',
                'data': detail_serializer.data
            }, status=status.HTTP_200_OK)

        print("ERRORS =", serializer.errors)

        return Response({
            'status': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk):
        return self.put(request, pk)
    

class MultipleAdmissionDeleteAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def delete(self, request):

        serializer = MultipleAdmissionDeleteSerializer(data=request.data)

        if serializer.is_valid():

            ids = serializer.validated_data['ids']

            deleted_count = Profiles.objects.filter(
                id__in=ids
            ).delete()

            return Response({
                'status': True,
                'message': 'Admissions deleted successfully',
                'deleted_count': deleted_count[0]
            }, status=status.HTTP_200_OK)

        return Response({
            'status': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    


    

# ==========================================
# FINANCE REPORT APIs
# ==========================================

class CourseReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        class_id = request.GET.get("class")

        classes = ClassModel.objects.filter(
    category=request.user.category
).order_by("class_name")

        if class_id:
            classes = classes.filter(id=class_id)

        report_data = []

        for class_obj in classes:

            academic_records = StudentAcademicDetails.objects.filter(
                student_class=class_obj,student_class__category=request.user.category
            )

            active_students = academic_records.filter(
                status__iexact="Active"
            ).count()

            completed_students = academic_records.filter(
                status__iexact="Completed"
            ).count()

            student_ids = academic_records.values_list(
                "user_id",
                flat=True
            )

            financials = StudentFinancialDetails.objects.filter(
                user_id__in=student_ids
            )

            revenue_generated = (
                financials.aggregate(
                    total=Sum("paid_amount")
                )["total"] or 0
            )

            pending_fees = (
                financials.aggregate(
                    total=Sum("balance_amount")
                )["total"] or 0
            )

            report_data.append({
                "id": class_obj.id,
                "course_standard": class_obj.class_name,
                "section" : class_obj.section ,
                "duration": ClassReportSerializer.calculate_duration(
                    class_obj.batch
                ),
                "active_students": active_students,
                "completed_students": completed_students,
                "revenue_generated": revenue_generated,
                "pending_fees": pending_fees,
                "total_teachers": (
                    1 if class_obj.class_teacher else 0
                ),
                "batch": class_obj.batch,
                "status": class_obj.status,
            })

        paginator = PageNumberPagination()
        paginator.page_size = int(
            request.GET.get("page_size", 10)
        )

        result_page = paginator.paginate_queryset(
            report_data,
            request
        )

        return paginator.get_paginated_response({
            "status": True,
            "data": result_page
        })


class StudentReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        queryset = Profiles.objects.filter(
            role='Student',
            studentacademicdetails__student_class__category=request.user.category
        ).select_related(
            'studentacademicdetails',
            'studentpersonaldetails',
            'studentfinancialdetails'
        ).order_by('-id')

        search = request.GET.get('search')

        if search:
            queryset = queryset.filter(
                Q(fullname__icontains=search) |
                Q(studentacademicdetails__admission_id__icontains=search) |
                Q(studentacademicdetails__student_class__class_name__icontains=search)
            )

        class_id = request.GET.get('class')

        if class_id:
            queryset = queryset.filter(
                studentacademicdetails__student_class_id=class_id
            )

        paginator = PageNumberPagination()
        paginator.page_size = int(
            request.GET.get('page_size', 10)
        )

        result_page = paginator.paginate_queryset(
            queryset,
            request
        )

        serializer = StudentReportSerializer(
            result_page,
            many=True
        )

        return paginator.get_paginated_response({
            'status': True,
            'data': serializer.data
        })

class TeacherReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        queryset = StaffManagementModel.objects.filter(
            is_teacher=True,
            profiles__category=request.user.category
        ).select_related(
            'profiles'
        ).order_by('-id')

        search = request.GET.get('search')

        if search:
            queryset = queryset.filter(
                Q(staff_name__icontains=search) |
                Q(profiles__user_id__icontains=search)
            )

        paginator = PageNumberPagination()
        paginator.page_size = int(
            request.GET.get('page_size', 10)
        )

        result_page = paginator.paginate_queryset(
            queryset,
            request
        )

        serializer = TeacherReportSerializer(
            result_page,
            many=True
        )

        return paginator.get_paginated_response({
            'status': True,
            'data': serializer.data
        })
    

    
from django.db.models.functions import ExtractMonth

class RevenueMonthListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        return Response({

            "status": True,

            "data": [

                {"id":1,"name":"January"},
                {"id":2,"name":"February"},
                {"id":3,"name":"March"},
                {"id":4,"name":"April"},
                {"id":5,"name":"May"},
                {"id":6,"name":"June"},
                {"id":7,"name":"July"},
                {"id":8,"name":"August"},
                {"id":9,"name":"September"},
                {"id":10,"name":"October"},
                {"id":11,"name":"November"},
                {"id":12,"name":"December"}

            ]
        })

class RevenueYearListAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        years = (
            StudentAcademicDetails.objects
            .dates(
                'admission_date',
                'year'
            )
        )

        return Response({
            "status": True,
            "data": [
                {
                    "year": year.year
                }
                for year in years
            ]
        })

class RevenueReportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        year = request.GET.get("year")
        month = request.GET.get("month")

        financial_queryset = StudentFinancialDetails.objects.all()

        if request.user.role != "SuperAdmin":
            financial_queryset = financial_queryset.filter(
                user__category=request.user.category
            )

        if year and year != "All Year":
            financial_queryset = financial_queryset.filter(
                user__studentacademicdetails__admission_date__year=int(year)
            )

        revenue_expression = ExpressionWrapper(
            F("admission_amount") +
            F("course_fee") -
            F("discount"),
            output_field=DecimalField()
        )

        total_admission_revenue = (
            financial_queryset.aggregate(
                total=Sum(revenue_expression)
            )["total"] or 0
        )

        total_fee_collection = (
            financial_queryset.aggregate(
                total=Sum("paid_amount")
            )["total"] or 0
        )

        total_pending_amount = (
            financial_queryset.aggregate(
                total=Sum("balance_amount")
            )["total"] or 0
        )

        staff_queryset = StaffManagementModel.objects.all()

        if request.user.role != "SuperAdmin":
            staff_queryset = staff_queryset.filter(
                profiles__category=request.user.category
            )

        total_salary_expense = (
            staff_queryset.aggregate(
                total=Sum("monthly_salary")
            )["total"] or 0
        )

        total_students = Profiles.objects.filter(
            role="Student"
        )

        total_teachers = StaffManagementModel.objects.filter(
            is_teacher=True
        )

        active_students = Profiles.objects.filter(
            role="Student",
            is_active=True
        )

        active_teachers = StaffManagementModel.objects.filter(
            is_teacher=True,
            profiles__is_active=True
        )

        if request.user.role != "SuperAdmin":

            total_students = total_students.filter(
                category=request.user.category
            )

            active_students = active_students.filter(
                category=request.user.category
            )

            total_teachers = total_teachers.filter(
                profiles__category=request.user.category
            )

            active_teachers = active_teachers.filter(
                profiles__category=request.user.category
            )

        total_students = total_students.count()
        total_teachers = total_teachers.count()
        active_students = active_students.count()
        active_teachers = active_teachers.count()

        monthly_data = []

        months_map = {
            1: "January",
            2: "February",
            3: "March",
            4: "April",
            5: "May",
            6: "June",
            7: "July",
            8: "August",
            9: "September",
            10: "October",
            11: "November",
            12: "December"
        }

        month_range = [int(month)] if month and month != "All Month" else range(1, 13)

        for month_number in month_range:

            month_financials = financial_queryset.filter(
                user__studentacademicdetails__admission_date__month=month_number
            )

            monthly_data.append({
                "month": month_number,
                "month_name": months_map[month_number],
                "admission_revenue": (
                    month_financials.aggregate(
                        total=Sum(revenue_expression)
                    )["total"] or 0
                ),
                "fee_collection": (
                    month_financials.aggregate(
                        total=Sum("paid_amount")
                    )["total"] or 0
                ),

                "pending_amount": (
                    month_financials.aggregate(
                        total=Sum("balance_amount")
                    )["total"] or 0
                ),
                "salary_expense": total_salary_expense
            })

        return Response({
            "status": True,
            "filters": {
                "year": year or "All Year",
                "month": month or "All Month"
            },
            "summary": {
                "total_admission_revenue": total_admission_revenue,
                "total_fee_collection": total_fee_collection,
                "total_pending_amount": total_pending_amount,
                "total_salary_expense": total_salary_expense,
                "total_students": total_students,
                "total_teachers": total_teachers,
                "active_students": active_students,
                "active_teachers": active_teachers
            },
            "chart_data": monthly_data
        }, status=status.HTTP_200_OK)





# class RevenueReportAPIView(APIView):

#     permission_classes = [IsAuthenticated]

#     def get(self, request):

#         year = request.GET.get("year")
#         month = request.GET.get("month")

#         financial_queryset = StudentFinancialDetails.objects.all()

#         if request.user.role != "SuperAdmin":
#             financial_queryset = financial_queryset.filter(
#                 user__category=request.user.category
#             )

#         if year and year != "All Year":
#             financial_queryset = financial_queryset.filter(
#                 user__studentacademicdetails__admission_date__year=int(year)
#             )

#         revenue_expression = ExpressionWrapper(
#             F("admission_amount") +
#             F("course_fee") -
#             F("discount"),
#             output_field=DecimalField()
#         )


#         total_admission_revenue = (
#             financial_queryset.aggregate(
#                 total=Sum(revenue_expression)
#             )["total"] or 0
#         )

#         total_fee_collection = (
#             financial_queryset.aggregate(
#                 total=Sum("paid_amount")
#             )["total"] or 0
#         )

#         total_pending_amount = (
#             financial_queryset.aggregate(
#                 total=Sum("balance_amount")
#             )["total"] or 0
#         )

#         staff_queryset = StaffManagementModel.objects.all()

#         if request.user.role != "SuperAdmin":
#             staff_queryset = staff_queryset.filter(
#                 profiles__category=request.user.category
#             )

#         total_salary_expense = (
#             staff_queryset.aggregate(
#                 total=Sum("monthly_salary")
#             )["total"] or 0
#         )

#         total_students = Profiles.objects.filter(
#             role="Student"
#         )

#         total_teachers = StaffManagementModel.objects.filter(
#             is_teacher=True
#         )

#         active_students = Profiles.objects.filter(
#             role="Student",
#             is_active=True
#         )

#         active_teachers = StaffManagementModel.objects.filter(
#             is_teacher=True,
#             profiles__is_active=True
#         )

#         if request.user.role != "SuperAdmin":

#             total_students = total_students.filter(
#                 category=request.user.category
#             )

#             active_students = active_students.filter(
#                 category=request.user.category
#             )

#             total_teachers = total_teachers.filter(
#                 profiles__category=request.user.category
#             )

#             active_teachers = active_teachers.filter(
#                 profiles__category=request.user.category
#             )

#         total_students = total_students.count()
#         total_teachers = total_teachers.count()
#         active_students = active_students.count()
#         active_teachers = active_teachers.count()

#         monthly_data = []

#         months_map = {
#             1: "January",
#             2: "February",
#             3: "March",
#             4: "April",
#             5: "May",
#             6: "June",
#             7: "July",
#             8: "August",
#             9: "September",
#             10: "October",
#             11: "November",
#             12: "December"
#         }

#         month_range = [int(month)] if month and month != "All Month" else range(1, 13)

#         for month_number in month_range:

#             month_financials = financial_queryset.filter(
#                 user__studentacademicdetails__admission_date__month=month_number
#             )

#             monthly_data.append({
#                 "month": month_number,
#                 "month_name": months_map[month_number],
#                 "admission_revenue": (
#                     month_financials.aggregate(
#                         total=Sum(revenue_expression)
#                     )["total"] or 0
#                 ),
#                 "fee_collection": (
#                     month_financials.aggregate(
#                         total=Sum("paid_amount")
#                     )["total"] or 0
#                 ),

#                 "pending_amount": (
#                     month_financials.aggregate(
#                         total=Sum("balance_amount")
#                     )["total"] or 0
#                 ),
#                 "salary_expense": total_salary_expense
#             })

#         return Response({
#             "status": True,
#             "filters": {
#                 "year": year or "All Year",
#                 "month": month or "All Month"
#             },
#             "summary": {
#                 "total_admission_revenue": total_admission_revenue,
#                 "total_fee_collection": total_fee_collection,
#                 "total_pending_amount": total_pending_amount,
#                 "total_salary_expense": total_salary_expense,
#                 "total_students": total_students,
#                 "total_teachers": total_teachers,
#                 "active_students": active_students,
#                 "active_teachers": active_teachers
#             },
#             "chart_data": monthly_data
#         }, status=status.HTTP_200_OK)
    



from django.shortcuts import render


def finance_test(request):
    return render(request, "test.html")


#EXPORTS 

class AdmissionExportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        print("ADMISSION EXPORT HIT")

        wb = Workbook()
        ws = wb.active
        ws.title = "Admissions"

        ws.append([
            "Admission ID",
            "Student Name",
            "Email",
            "Course",
            "Admission Date",
            "Admission Amount",
            "Paid Amount",
            "Balance Amount",
            "Payment Mode",
            "Payment Status"
        ])

        queryset = StudentFinancialDetails.objects.select_related(
            'user'
        )

        for obj in queryset:

            try:
                academic = obj.user.studentacademicdetails
                course = (
                    academic.courses.name
                    if academic.courses else ''
                )
            except:
                academic = None
                course = ''

            payment_status = (
                "Paid"
                if obj.balance_amount <= 0
                else "Partial"
            )

            ws.append([
                academic.admission_id if academic else '',
                obj.user.fullname if obj.user else '',
                obj.user.email if obj.user else '',
                course,
                academic.admission_date if academic else '',
                obj.admission_amount,
                obj.paid_amount,
                obj.balance_amount,
                obj.payment_mode,
                payment_status
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename=admissions_report.xlsx'
        )

        wb.save(response)

        return response
    
class StudentReportExportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"

        ws.append([
            "Admission ID",
            "Student Name",
            "Phone Number",
            "Parent Number",
            "Course",
            "Batch",
            "Gender",
            "Collected Fees",
            "Pending Fees",
            "Attendance Percentage",
            "Status"

        ])

        students = Profiles.objects.filter(
            role='Student'
        )

        for student in students:

            try:
                academic = student.studentacademicdetails
                admission_id = academic.admission_id
                course = academic.courses.name if academic.courses else ''
                batch = academic.batch
            except:
                admission_id = ''
                course = ''
                batch = ''

            try:
                personal = student.studentpersonaldetails

                gender = personal.gender
                parent_number = (
                    personal.parent_guardian_phone_number
                )
            except:
                gender = ''
                parent_number = ''

            try:
                financial = student.studentfinancialdetails

                collected_fees = financial.paid_amount
                pending_fees = financial.balance_amount
            except:
                collected_fees = 0
                pending_fees = 0

            # Replace later with actual attendance model
            attendance_percentage = 'N/A'

            ws.append([
                admission_id,
                student.fullname,
                student.phone_number,
                parent_number,
                course,
                batch,
                gender,
                collected_fees,
                pending_fees,
                attendance_percentage,
                (
                    "Active"
                    if student.is_active
                    else "Inactive"
                )
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename=student_report.xlsx'
        )

        wb.save(response)

        return response
    
class TeacherReportExportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        wb = Workbook()
        ws = wb.active
        ws.title = "Teachers"

        ws.append([
            "Teacher ID",
            "Teacher Name",
            "Email",
            "Phone Number",
            "Joining Date",
            "Designation",
            "Department",
            "Qualification",
            "Experience",
            "Incharge Class",
            "Salary",
            "Status"
        ])

        teachers = StaffManagementModel.objects.filter(
            is_teacher=True
        ).select_related('profiles')

        for teacher in teachers:

            ws.append([
                teacher.profiles.user_id
                if teacher.profiles else '',

                teacher.staff_name,

                teacher.profiles.email
                if teacher.profiles else '',

                teacher.profiles.phone_number
                if teacher.profiles else '',

                teacher.joining_date,

                teacher.designation,

                teacher.department,

                teacher.qualification,

                teacher.experience_year,

                teacher.student_class.class_name
                if teacher.student_class else '',

                teacher.monthly_salary,

                (
                    "Active"
                    if teacher.profiles
                    and teacher.profiles.is_active
                    else "Inactive"
                )
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename=teacher_report.xlsx'
        )

        wb.save(response)

        return response
    

class RevenueReportExportAPIView(APIView):

    permission_classes = [IsAuthenticated]

    def get(self, request):

        wb = Workbook()
        ws = wb.active
        ws.title = "Revenue"

        revenue_expression = ExpressionWrapper(
            F("admission_amount") +
            F("course_fee") -
            F("discount"),
            output_field=DecimalField()
        )

        financial_queryset = StudentFinancialDetails.objects.all()
        staff_queryset = StaffManagementModel.objects.all()
        student_queryset = Profiles.objects.filter(role='Student')
        teacher_queryset = StaffManagementModel.objects.filter(is_teacher=True)

        if request.user.role != 'SuperAdmin':
            financial_queryset = financial_queryset.filter(
                user__category=request.user.category
            )
            staff_queryset = staff_queryset.filter(
                profiles__category=request.user.category
            )
            student_queryset = student_queryset.filter(
                category=request.user.category
            )
            teacher_queryset = teacher_queryset.filter(
                profiles__category=request.user.category
            )

        total_admission_revenue = (
            StudentFinancialDetails.objects.aggregate(
                total=Sum(revenue_expression)
            )['total'] or 0
        )

        total_fee_collection = (
            StudentFinancialDetails.objects.aggregate(
                total=Sum('paid_amount')
            )['total'] or 0
        )

        total_pending_amount = (
            StudentFinancialDetails.objects.aggregate(
                total=Sum('balance_amount')
            )['total'] or 0
        )

        total_salary_expense = (
            StaffManagementModel.objects.aggregate(
                total=Sum('monthly_salary')
            )['total'] or 0
        )

        total_students = Profiles.objects.filter(
            role='Student'
        ).count()

        total_teachers = StaffManagementModel.objects.filter(
            is_teacher=True
        ).count()

        active_students = Profiles.objects.filter(
            role='Student',
            is_active=True
        ).count()

        active_teachers = StaffManagementModel.objects.filter(
            is_teacher=True,
            profiles__is_active=True
        ).count()

        net_profit = (
            total_fee_collection -
            total_salary_expense
        )

        ws.append(["Metric", "Amount"])

        ws.append([
            "Total Admission Revenue",
            total_admission_revenue
        ])

        ws.append([
            "Total Fee Collection",
            total_fee_collection
        ])

        ws.append([
            "Total Pending Amount",
            total_pending_amount
        ])

        ws.append([
            "Total Salary Expense",
            total_salary_expense
        ])

        ws.append([
            "Net Profit",
            net_profit
        ])

        ws.append([
            "Total Students",
            total_students
        ])

        ws.append([
            "Total Teachers",
            total_teachers
        ])

        ws.append([
            "Active Students",
            active_students
        ])

        ws.append([
            "Active Teachers",
            active_teachers
        ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename=revenue_report.xlsx'
        )

        wb.save(response)

        return response